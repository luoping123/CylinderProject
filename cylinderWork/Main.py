import math
import random
import re
import sys
import threading
import time
from datetime import datetime, timedelta
import matplotlib.dates as mdate

import mplcursors
import numpy as np
import openpyxl
import pandas as pd
from PySide2.QtWidgets import QApplication, QWidget, QFileDialog, QTableWidgetItem, QAbstractItemView, QHeaderView, \
    QTableWidget, QFrame, QVBoxLayout, QMessageBox
from PySide2.QtUiTools import QUiLoader
from PySide2.QtCore import QFile, QIODevice, Qt, QObject, Signal
from matplotlib.animation import FuncAnimation
from matplotlib.ticker import AutoLocator, MultipleLocator, MaxNLocator
from qt_material import apply_stylesheet
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from scipy.optimize import curve_fit
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split
import pyvista as pv
# from pyvistaqt import BackgroundPlotter
from pyvista.plotting.opts import PickerType
import os
os.environ["QT_API"] = "pyside2"
from pyvistaqt import QtInteractor, BackgroundPlotter, MultiPlotter
# TODO 导入自己写的模块，vtu模型拟合和生成vtu文件
from BDFAndTemperatureToVtuText7 import VtuModel, strain_stress_to_vtu_1HZ, get_rtz_from_excel, eFunFitting_to_vtuModel
from ViewPoint import ViewPoints    # TODO 自定义监测点类
from CylinderCustomTool import CylinderCustomTool   # TODO 自定义的圆柱模型尺寸及网格疏密修改工具类


'''在13的基础上修改，改变UI为main5，增加模型初始化设置，功能实现目标：监测点位置 从模板Excel文件中读取配置'''

plt.rcParams["font.family"] = "Microsoft YaHei"
# pv.global_theme.font.family = 'arial'
# pv.set_plot_theme('paraview')


data_excel = None   # Excel数据，每秒增加一行
test_Temperature_lists =None    # 监测点温度列表 二维，用于拟合圆柱温度
bg_Temperature_list = None  # 环境温度列表，用于判断何时开始降温
time_list = None    # 时间列表 一维
real_time_list = None   #  真实时间列表，年月日 时分秒
strain_lists = None # 应变列表 二维
start_time_of_decrease = 0  # 开始降温时刻
decrease_index = 0  # 开始降温时刻索引，用于在整个列表定位
end_time = 0    # 当前实验时刻，[降温开始, end_time]用于预测训练，每秒变化
target_file_path = None #
vtu_storage_directory = None # 模型文件存储文件夹，用于历史云图追溯
# 已知的温度 监测点的坐标和数据
# known_points = get_rtz_from_excel() # TODO BDFAndTemperatureToVtuText5.py的方法
# bdf_filename = "../CylinderData/CylinderDemo5.bdf"  # TODO 后期存储到配置文件中，文件必须存在
now_vtu_filename = "../CylinderData/cylinderdemo5_out"  # 当前时刻vtu文件名 临时的，后期会换掉
# template_vtu_filename = "../CylinderData/cylinderdemo5_out.vtu"    # TODO 后期存储到配置文件中，文件必须存在
history_vtu_filename = None # 云图追溯的vtu文件，通过选中表格 某行进行动态修改
# 把bdf文件中的节点、单元信息全部设置到vtuModel对象中
# vtuModel = VtuModel(bdf_filename, now_vtu_filename + ".vtu")
vtuModel = VtuModel()
TIME_TOONE = 0 # TODO 开始降温时刻的 相对时间（Times列），用于时间的归一处理
STRAIN_LIST_TOONE = None    # TODO 开始降温时刻的  应变列表首行，用于应变的归一处理
view_points_dic = {}    # 监测点字典，包含监测点
# TODO 用于Matplotlib缩放
startx = 0
starty = 0
mPress = False
# TODO 鼠标是否在Matplotlib的Axis上，若不在会跟新坐标轴，来保持图像居中
mouse_in_axis = False
ax3_strain_legend_2 = None   #

view_points_mesh_dic = {}   # key:english_column_name，value:mesh
view_points_labels_dic = {}     # key:english_column_name，value:标注的坐标
view_points_coordinates_dic = {}    # key:english_column_name，value:监测点的坐标
point_label_actor = None    # 实时监测点的标注
history_vtu_directory = None    #  历史云图的目录
pickPoint = None        #  鼠标选中的点 [x,y,z]
pickPoint_index = None      # 鼠标选中的点的索引




class MySignals(QObject):
    # 定义一种信号
    # 调用 emit方法发信号时，传入参数必须是这里指定的参数类型
    text_print = Signal(np.ndarray, list)
    pickPointLabel = Signal(pv.core.pyvista_ndarray, np.int64)


class MyWindow:
    def __init__(self, fig):
        # 从文件中加载UI定义
        self.load_ui()
        # global template_vtu_filename, known_points, axs
        global axs
        # vtu_filename = "../Data/cylinderdemo4_out.vtu"
        # 读取 VTK 文件
        # mesh = pv.read(template_vtu_filename)
        # mesh.set_active_scalars("Temperature")
        # self.mesh = mesh
        # self.mesh_slice = mesh.copy()
        # self.mesh_now = mesh.copy()

        self.ms = MySignals()                              #引入信号函数
        self.ms.text_print.connect(self.update_plotter_now_labels)           #将信号传递给主程序中pF函数进行处理
        self.ms.pickPointLabel.connect(self.update_pickPoint_label)     # 将信号传递给主程序update_pickPoint_label函数处理

        cylinderStructured = pv.CylinderStructured(radius=np.linspace(0, 55, 10), height=110,
                                                   direction=(0.0, 0.0, 1.0), theta_resolution=30, z_resolution=30)
        # 将结构化网格转换为非结构化四面体网格
        self.cylinder_mesh = cylinderStructured.triangulate()
        self.mesh = self.cylinder_mesh.copy()
        self.mesh_slice = self.cylinder_mesh.copy()
        self.mesh_now = self.cylinder_mesh.copy()


        # 给选择模型存放文件夹按钮 绑定单击事件
        self.ui.selectFDirectoryButton.clicked.connect(self.click_openDirectory)
        # 给选择文件按钮 绑定单击事件
        self.ui.selectFileButton.clicked.connect(self.click_openFile)
        # 给选择监测点配置文件 绑定单击事件
        self.ui.viewPointButton.clicked.connect(self.click_openFileOfViewPoint)

        # 给测试按钮 绑定单击事件
        # self.ui.pushButton_start.clicked.connect(self.click_start)

        # 给表格自动滚动 选择框 绑定更改事件
        # self.ui.checkBox_tableAutoScroll.stateChanged.connect(self.click_tableAutoScroll)

        # 给选行追溯云图 选择框 绑定更改事件
        # self.ui.checkBox_tableSelectEnable.stateChanged.connect(self.click_tableSelectEnable)

        # 给保存模型  按钮 绑定单击事件
        self.ui.btn_saveModel.clicked.connect(self.click_btn_saveModel)

        # 给重置模型 按钮 绑定单击事件
        self.ui.btn_resetModel.clicked.connect(self.click_btn_resetModel)

        # 给 选择历史云图存放文件夹 按钮 绑定单击事件
        self.ui.btn_selectCloudDirectory.clicked.connect(self.click_selectHistoryVtuDirectory)

        # 给 选择历史实验数据 按钮 绑定单击事件
        self.ui.btn_selectHistoryData.clicked.connect(self.click_selectHistoryDataFile)

        self.fig = fig
        self.canvas = FigureCanvas(fig)
        # TODO 预测曲线添加
        self.ui.hLayout_prePlot.addWidget(self.canvas)

        # TODO 设置Matplotlib 曲线图
        self.ax_temperature1 = axs[0, 0]
        self.ax_temperature2 = axs[0, 1]
        self.ax_strain1 = axs[1, 0]
        self.ax_strain2 = axs[1, 1]

        self.ax_temperature1_lines = {}
        self.ax_temperature2_lines = {}
        self.ax_strain1_lines = {}
        self.ax_strain2_lines = {}



        # self.renderer = self.get_Pyvista_renderer()

        # self.frame = QFrame()
        # vlayout = QVBoxLayout()

        # add the pyvista interactor object
        # self.plotter = QtInteractor(self.frame)
        # TODO 模型尺寸和网格疏密交互式修改
        # self.plotter_now = BackgroundPlotter(toolbar=False, menu_bar=False, update_app_icon=True)
        self.plotter_modelModify = MultiPlotter(nrows=1, ncols=1, title='模型设定',
                                        auto_update=False,
                                        toolbar=False, menu_bar=False, update_app_icon=True)
        self.plotter_modelModify[0, 0].set_background('black', top='grey')  # TODO 背景颜色白黑渐变  grey  white
        # cylinderStructured = pv.CylinderStructured(radius=np.linspace(0, 55, 100), height=2.0,
        #                                            direction=(0.0, 0.0, 1.0), theta_resolution=30, z_resolution=30)
        # 将结构化网格转换为非结构化四面体网格
        # self.cylinder_mesh = cylinderStructured.triangulate()
        # 配置尺寸自定义工具
        self.modelCustomTool = CylinderCustomTool(self.cylinder_mesh)

        self.modelModify_actor = self.plotter_modelModify[0, 0].add_mesh(self.cylinder_mesh, name="modelModify", show_edges=True,
                                      # style="points",
                                      )
        self.radius_slider = self.plotter_modelModify[0, 0].add_slider_widget(
            callback=lambda value: self.modelCustomTool('radius_lenth', value),
            rng=[30, 100],
            value=55,
            title="radius(mm)",
            pointa=(0.03, 0.6),
            pointb=(0.25, 0.6),
            fmt='%0.f',  # 设置滑块的值为整数
            style='modern',
            color='white',
        )
        self.height_slider = self.plotter_modelModify[0, 0].add_slider_widget(
            callback=lambda value: self.modelCustomTool('height', value),
            rng=[30, 200],
            value=110,
            title="height(mm)",
            pointa=(0.03, 0.4),
            pointb=(0.25, 0.4),
            fmt='%0.f',  # 设置滑块的值为整数
            style='modern',
            color='white',
        )
        self.r_resolution_slider = self.plotter_modelModify[0, 0].add_slider_widget(
            callback=lambda value: self.modelCustomTool('radius_resolution', value),
            rng=[5, 15],
            value=10,
            title="r_resolution(level)",
            pointa=(0.75, 0.7),
            pointb=(0.97, 0.7),
            fmt='%0.f',  # 设置滑块的值为整数
            style='modern',
            color='white',
        )
        self.theta_resolution_slider = self.plotter_modelModify[0, 0].add_slider_widget(
            callback=lambda value: self.modelCustomTool('theta_resolution', int(value)),
            rng=[50, 200],
            value=100,
            title="theta_resolution(level)",
            pointa=(0.75, 0.5),
            pointb=(0.97, 0.5),
            fmt='%0.f',  # 设置滑块的值为整数
            style='modern',
            color='white',
        )
        self.z_resolution_slider = self.plotter_modelModify[0, 0].add_slider_widget(
            callback=lambda value: self.modelCustomTool('z_resolution', int(value)),
            rng=[10, 20],
            value=15,
            title="z_resolution(level)",
            pointa=(0.75, 0.3),
            pointb=(0.97, 0.3),
            fmt='%0.f',  # 设置滑块的值为整数
            style='modern',
            color='white',
        )
        # 设置滑块的颜色
        self.radius_slider.GetRepresentation().GetSliderProperty().SetColor((0.26667, 0.5411, 1))
        self.height_slider.GetRepresentation().GetSliderProperty().SetColor((0.26667, 0.5411, 1))
        self.r_resolution_slider.GetRepresentation().GetSliderProperty().SetColor((0.26667, 0.5411, 1))
        self.theta_resolution_slider.GetRepresentation().GetSliderProperty().SetColor((0.26667, 0.5411, 1))
        self.z_resolution_slider.GetRepresentation().GetSliderProperty().SetColor((0.26667, 0.5411, 1))


        self.plotter_modelModify[0, 0].add_axes(  # TODO 左下角坐标轴
            line_width=5,
            color='white',
            cone_radius=0.6,
            shaft_length=0.7,
            tip_length=0.3,
            ambient=0.5,
            label_size=(0.4, 0.16),
        )
        self.plotter_modelModify[0, 0].add_camera_orientation_widget()  # TODO 右上角坐标轴带负轴
        # TODO 实时模型显示
        # self.plotter_now = BackgroundPlotter(toolbar=False, menu_bar=False, update_app_icon=True)
        self.plotter_now = MultiPlotter(nrows=1, ncols=1, title='实时云图',
                                        auto_update=False,
                                        toolbar=False, menu_bar=False, update_app_icon=True)
        self.plotter_now[0,0].set_background('black', top='grey') # TODO 背景颜色白黑渐变  grey  white
        # TODO 历史模型追溯和切片
        self.plotter = MultiPlotter(nrows=1, ncols=2, toolbar=False, menu_bar=False, update_app_icon=True)
        self.plotter[0, 0].set_background('black', top='grey') # TODO 背景颜色白黑渐变
        self.plotter[0, 1].set_background('black', top='grey') # TODO 背景颜色白黑渐变
        self.model_slice = self.plotter[0, 1].add_mesh(self.mesh_slice, name="model_slice", cmap="coolwarm",
                                                 show_scalar_bar=False,
                                                 copy_mesh=True,
                                                 point_size=5,  # TODO   点的大小
                                                 style='surface',
                                                 # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
                                                 )  # TODO   圆柱
        self.model_slice.visibility = False
        # TODO 空的云图
        self.model_none = self.plotter_now[0,0].add_mesh(self.mesh, name="model_none",)

        # TODO 当前时刻的云图
        self.mesh_now.point_data["Temperature"] = np.array([0] * len(self.mesh_now.points))
        self.mesh_now.set_active_scalars("Temperature")
        self.model_now = self.plotter_now[0,0].add_mesh(self.mesh_now, name="model_now", cmap="coolwarm",
                                  show_scalar_bar=False,
                                  # show_vertices=True,  # TODO   显示点
                                  point_size=5,  # TODO   点的大小
                                  style='surface',
                                  # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
                                  )  # TODO   圆柱
        self.model_now.visibility = False

        self.plotter_now[0,0].add_axes(  # TODO 左下角坐标轴
            line_width=5,
            color='white',
            cone_radius=0.6,
            shaft_length=0.7,
            tip_length=0.3,
            ambient=0.5,
            label_size=(0.4, 0.16),
        )
        self.plotter_now[0,0].add_camera_orientation_widget()  # TODO 右上角坐标轴带负轴

        # TODO 在plotter_now[0,0]中绘制曲线
        self.temperature_chart = pv.Chart2D(size=(0.45, 0.30), loc=(0.53, 0.40), x_label="Time (s)", y_label="Temperature")
        # self.temperature_lines1 = {}
        self.temperature_history_line = self.temperature_chart.line([0], [0], color='red', label='T_history',
                                                                    style='-')  # 初始为空数据
        self.temperature_future_line = self.temperature_chart.line([0], [0], color='red', label='T_future',
                                                                    style='--')  # 初始为空数据
        self.temperature_chart.background_color = (1.0, 1.0, 1.0, 0.4)
        self.temperature_chart.legend_visible = False
        self.temperature_chart_actor = self.plotter_now[0,0].add_chart(self.temperature_chart)
        self.temperature_chart.visible = False     # 隐藏绘图

        self.strain_chart = pv.Chart2D(size=(0.45, 0.30), loc=(0.53, 0.03), x_label="Time (s)",
                                            y_label="strain")
        # self.temperature_lines2 = {}
        self.r_strain_history_line = self.strain_chart.line([0], [0], color='green', label='r_history',
                                                                    style='-')  # 初始为空数据
        self.r_strain_future_line = self.strain_chart.line([0], [0], color='green', label='r_future',
                                                          style='--')  # 初始为空数据
        self.theta_strain_history_line = self.strain_chart.line([0], [0], color='yellow', label='t_history',
                                                            style='-')  # 初始为空数据
        self.theta_strain_future_line = self.strain_chart.line([0], [0], color='yellow', label='t_future',
                                                           style='--')  # 初始为空数据
        self.z_strain_history_line = self.strain_chart.line([0], [0], color='blue', label='z_history',
                                                                style='-')  # 初始为空数据
        self.z_strain_future_line = self.strain_chart.line([0], [0], color='blue', label='z_future',
                                                               style='--')  # 初始为空数据
        self.strain_chart.background_color = (1.0, 1.0, 1.0, 0.4)
        self.strain_chart.legend_visible = False
        self.strain_chart_actor = self.plotter_now[0, 0].add_chart(self.strain_chart)
        self.strain_chart.visible = False     # 隐藏绘图

        # self.plotter.set_background('grey', top='white')
        # self.plotter.add_editor()
        # 把PySide嵌入到界面中
        # self.ui.vLayout_model.addWidget(self.plotter.interactor)
        # TODO 模型设定修改
        self.ui.vLayout_modelmodify.addWidget(self.plotter_modelModify._window)
        # TODO 实时模型显示
        # self.ui.hLayout_preModel.addWidget(self.plotter_now.app_window)
        self.ui.hLayout_preModel.addWidget(self.plotter_now._window)
        # TODO 历史模型追溯和切片
        self.ui.vLayout_model.addWidget(self.plotter._window)
        # vlayout.addWidget(self.plotter.interactor)
        # self.ui.signal_close.connect(self.plotter.close)

        # self.frame.setLayout(vlayout)
        # 显示模型
        self.ui.btn_showModel.clicked.connect(self.model_show)
        # 切片显示
        self.ui.btn_showSliceModel.clicked.connect(self.slice_model_show)
        # 数据类型
        self.ui.cbox_styleType.currentIndexChanged.connect(self.handleSelectionChange)

        # 把PySide嵌入到界面中
        # self.ui.vLayout_model.addWidget(self.renderer)


        # 设置表格宽度自适应
        # self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # 清除表格内容和表头
        self.ui.tableWidget.clear()
        # TODO 优化3 将表格变为禁止编辑
        self.ui.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ui.tableWidget_History.setEditTriggers(QAbstractItemView.NoEditTriggers)
        #TODO 优化 4 设置表格整行选中
        self.ui.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ui.tableWidget_History.setSelectionBehavior(QAbstractItemView.SelectRows)
        # TODO 给表格添加选中 触发事件
        self.ui.tableWidget_History.itemSelectionChanged.connect(self.table_history_selected_item)
        # TODO 设置表格不可选中
        self.ui.tableWidget_History.setSelectionMode(QAbstractItemView.SingleSelection)
        # 设置背景色交替
        # self.ui.tableWidget.setAlternatingRowColors(True)
        # self.ui.tableWidget.setStyleSheet("alternate-background-color: lightyellow; background-color: white ;")
        # TODO
        # vtu_filename2= "../CylinderData/Text1/CylinderModelTime-87-FittingError.vtu"

        # self.update_plotter_now(vtu_filename2)

        # self.ui.gridLayout.add

    def load_ui(self):
        # 从文件中加载UI定义
        ui_file_name = 'UI/main5.ui'
        ui_file = QFile(ui_file_name)
        if not ui_file.open(QIODevice.ReadOnly):
            print(f"Cannot open {ui_file_name}: {ui_file.errorString()}")
            sys.exit(-1)
        # 从 UI 定义中动态 创建一个相应的窗口对象
        # 注意：里面的控件对象也成为窗口对象的属性了
        # 比如 self.ui.button , self.ui.textEdit
        loader = QUiLoader()
        self.ui = loader.load(ui_file)
        ui_file.close()
        if not self.ui:
            print(loader.errorString())
            sys.exit(-1)

    # 云图显示
    def model_show(self):
        global history_vtu_filename
        if history_vtu_filename is None or history_vtu_filename == '':
            # 显示警告弹窗
            QMessageBox.warning(self.ui, '警告', '请先选择表格某行数据！')
            return
        # vtu_filename = "../Data/cylinderdemo4_out.vtu"
        # 读取 VTK 文件
        print(history_vtu_filename)
        history_mesh = pv.read(history_vtu_filename)
        # 读取节点数据
        node_data_name = "Temperature"  # 替换为节点数据名称  默认温度
        node_data_name = self.ui.cbox_dataType.currentText()
        node_data = history_mesh.point_data[node_data_name] # TODO 将所选中的vtu文件 数据取出 动态设置数据
        # self.mesh.set_active_scalars(node_data_name)    # TODO 设置当前活跃的 属性
        style = self.ui.cbox_styleType.currentText()

        # 清除所有东西
        self.plotter[0,0].clear()
        self.plotter[0,0].clear_plane_widgets()
        # 重新设置 默认灯光
        self.plotter[0,0].enable_lightkit()

        scalar_bars = self.plotter[0,0].scalar_bars.values()
        scalar_bar_list = list(scalar_bars)
        for scalar_bar in scalar_bar_list:
            self.plotter[0,0].remove_scalar_bar()
        # self.plotter.remove_scalar_bar()

        self.plotter[0,0].disable_picking()
        # TODO 点击 点 后显示其值
        def fun_pick(point):
            points = self.mesh.points
            # 计算点与数组中每个点的距离
            distances = np.linalg.norm(points - point, axis=1)
            # 找到距离最接近的点的索引
            index = np.argmin(distances)
            self.plotter[0,0].add_point_labels(np.array([points[index]]), np.array([node_data[index]]),
                                          point_color='black',
                                          point_size=10,
                                          font_size=20,
                                          name='point_1'  # TODO 添加的参与者的名称，以便于更新。如果此名称的参与者已经存在于呈现窗口中，它将被新参与者替换。
                                          )

        if style == 'surface with edges':
            # self.surface_edges.visibility = True
            self.plotter[0, 0].add_mesh(self.mesh, name="model", scalars=node_data, cmap="coolwarm",
                                        # show_scalar_bar=False,
                                        scalar_bar_args={'title': node_data_name,
                                                         'color': 'white',
                                                         'label_font_size': 14,
                                                         "vertical": "False",
                                                         "height": 0.65,
                                                         "position_x": 0.85,
                                                         "position_y": 0.1},
                                        # show_vertices=True,  # TODO   显示点
                                        point_size=3,  # TODO   点的大小
                                        # style='points_gaussian',
                                        show_edges=True,  # TODO   显示网格
                                        style='surface',
                                        # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
                                        )  # TODO   圆柱
        elif style == 'gauss point':
            self.plotter[0, 0].add_mesh(pv.PolyData(self.mesh.points), name="model", scalars=node_data, cmap='coolwarm',
                                        pickable=True,
                                        # show_scalar_bar=False,
                                        scalar_bar_args={'title': node_data_name,
                                                         'color': 'white',
                                                         'label_font_size': 14,
                                                         "vertical": "False",
                                                         "height": 0.65,
                                                         "position_x": 0.85,
                                                         "position_y": 0.1},
                                        # show_vertices=True,  # TODO   显示点
                                        point_size=5,  # TODO   点的大小
                                        # style='points_gaussian',
                                        show_edges=True,  # TODO   显示网格
                                        style='points',
                                        # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
                                        )

            # TODO 拾取点
            self.plotter[0, 0].enable_point_picking(callback=fun_pick,
                                                    # TODO 回调函数，第一个参数是被点击的点坐标 [-33.262 -27.3781 -22.936]
                                                    # tolerance=0.025,  # TODO 拾取的误差，按照屏幕百分比
                                                    # left_clicking=False,  # TODO 默认鼠标右键选取，True则鼠标左键
                                                    # picker=PickerType.POINT,  # TODO 选取的类型 hardware、cell、point、volume
                                                    # show_message=True,  # TODO 显示关于如何使用点拾取工具的消息。如果这是一个字符串，那将是显示的消息。
                                                    # font_size=18,  # TODO 设置消息的大小。
                                                    # color='black',  # TODO 显示被选中后的颜色。
                                                    # point_size=10,  # TODO 如果' show_point ' '为' True'，则所选点的大小。
                                                    # show_point=True,  # TODO  点击后显示选中点。
                                                    # use_picker=False,  # TODO  当' ' True ' '时，回调函数也将被传递给选择器。
                                                    # pickable_window=False,  # TODO 当“True”并且所选拾取器支持它时，3D窗口中的点是可拾取的。
                                                    # clear_on_no_selection=True,  # TODO 当没有选择任何点时清除所选内容。
                                                    # **kwargs,  # TODO
                                                    )
        elif style == 'points':
            self.plotter[0, 0].add_mesh(self.mesh, name="model", scalars=node_data, cmap="coolwarm",
                                        # show_scalar_bar=False,
                                        scalar_bar_args={'title': node_data_name,
                                                         'color': 'white',
                                                         'label_font_size': 14,
                                                         "vertical": "False",
                                                         "height": 0.65,
                                                         "position_x": 0.85,
                                                         "position_y": 0.1},
                                        # show_vertices=True,  # TODO   显示点
                                        point_size=5,  # TODO   点的大小
                                        # style='points_gaussian',
                                        # show_edges=True,     # TODO   显示网格
                                        style=style,
                                        # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
                                        )  # TODO   圆柱

            # TODO 拾取点
            self.plotter[0, 0].enable_point_picking(callback=fun_pick,
                                                    # TODO 回调函数，第一个参数是被点击的点坐标 [-33.262 -27.3781 -22.936]
                                                    # tolerance=0.025,  # TODO 拾取的误差，按照屏幕百分比
                                                    # left_clicking=False,  # TODO 默认鼠标右键选取，True则鼠标左键
                                                    # picker=PickerType.POINT,  # TODO 选取的类型 hardware、cell、point、volume
                                                    # show_message=True,  # TODO 显示关于如何使用点拾取工具的消息。如果这是一个字符串，那将是显示的消息。
                                                    # font_size=18,  # TODO 设置消息的大小。
                                                    # color='black',  # TODO 显示被选中后的颜色。
                                                    # point_size=10,  # TODO 如果' show_point ' '为' True'，则所选点的大小。
                                                    # show_point=True,  # TODO  点击后显示选中点。
                                                    # use_picker=False,  # TODO  当' ' True ' '时，回调函数也将被传递给选择器。
                                                    # pickable_window=False,  # TODO 当“True”并且所选拾取器支持它时，3D窗口中的点是可拾取的。
                                                    # clear_on_no_selection=True,  # TODO 当没有选择任何点时清除所选内容。
                                                    # **kwargs,  # TODO
                                                    )
        else:
            self.plotter[0, 0].add_mesh(self.mesh, name="model", scalars=node_data, cmap="coolwarm",
                                                     # show_scalar_bar=False,
                                                     scalar_bar_args={'title': node_data_name,
                                                                      'color': 'white',
                                                                      'label_font_size': 14,
                                                                      "vertical": "False",
                                                                      "height": 0.65,
                                                                      "position_x": 0.85,
                                                                      "position_y": 0.1},
                                                     # show_vertices=True,  # TODO   显示点
                                                     point_size=5,  # TODO   点的大小
                                                     # style='points_gaussian',
                                                     # show_edges=True,     # TODO   显示网格
                                                     style=style,
                                                     # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
                                                     )  # TODO   圆柱



        self.plotter[0,0].add_axes(  # TODO 左下角坐标轴
            line_width=5,
            color='white',
            cone_radius=0.6,
            shaft_length=0.7,
            tip_length=0.3,
            ambient=0.5,
            label_size=(0.4, 0.16),
        )

        self.plotter[0,0].add_camera_orientation_widget()  # TODO 右上角坐标轴带负轴

        # plotter = pv.Plotter()
        # self.plotter.add_mesh(pv.PolyData(mesh.points), scalars=node_data, cmap='coolwarm', pickable=True, style='points')




        self.plotter[0,0].reset_camera()
        # 获取Pyvista的渲染窗口
        # renderer = plotter.show()
        # print(renderer)
        # return renderer

    # 云图切片
    def slice_model_show(self):
        global history_vtu_filename
        global history_vtu_filename
        if history_vtu_filename is None or history_vtu_filename == '':
            # 显示警告弹窗
            QMessageBox.warning(self.ui, '警告', '请先选择表格某行数据！')
            return
        sliceType = self.ui.cbox_sliceType.currentText()

        if sliceType == '无':
            # 显示警告弹窗
            QMessageBox.warning(self.ui, '警告', '请先将显示类型调整为Surface！')
            return
        # vtu_filename = "../Data/cylinderdemo4_out.vtu"
        # 读取 VTK 文件
        history_mesh = pv.read(history_vtu_filename)
        # 读取节点数据
        node_data_name = "Temperature"  # 替换为节点数据名称  默认温度
        node_data_name = self.ui.cbox_dataType.currentText()
        node_data = history_mesh.point_data[node_data_name]
        self.mesh_slice.point_data[node_data_name] = node_data
        self.mesh_slice.set_active_scalars(node_data_name)    # TODO 设置当前活跃的 属性
        self.model_slice.visibility = True    # TODO 显示actor
        # scalar_bar = self.plotter[0, 1].scalar_bar
        # print("scalar_bar=======>>>>>")
        # print(scalar_bar)
        # 清除所有东西
        # self.plotter[0, 1].clear()
        box_widgets_list = self.plotter[0, 1].box_widgets
        # print(self.plotter[0, 1].box_widgets)
        for box in box_widgets_list:
            box.Off()
        self.plotter[0, 1].clear_slider_widgets()
        self.plotter[0, 1].remove_actor("model_slice2")
        # self.plotter[0, 1].clear_box_widgets()
        # 重新设置 默认灯光
        # self.plotter[0, 1].enable_lightkit()

        scalar_bars = self.plotter[0, 1].scalar_bars.values()
        scalar_bar_list = list(scalar_bars)
        for scalar_bar in scalar_bar_list:
            self.plotter[0, 1].remove_scalar_bar()
        # self.plotter.remove_scalar_bar()

        # self.plotter[0, 0].disable_picking()

        # self.model = self.plotter[0, 1].add_mesh(mesh, name="model", scalars=node_data, cmap="coolwarm",
        #                             # show_scalar_bar=False,
        #                             scalar_bar_args={'title': node_data_name,
        #                                              "vertical": "False",
        #                                              "height": 0.65,
        #                                              "position_x": 0.85,
        #                                              "position_y": 0.1},
        #                             # show_vertices=True,  # TODO   显示点
        #                             point_size=5,  # TODO   点的大小
        #                             # style='points_gaussian',
        #                             # show_edges=True,     # TODO   显示网格
        #                             style='surface',
        #                             # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
        #                             )  # TODO   圆柱



        sliceType = self.ui.cbox_sliceType.currentText()

        if sliceType == '盒子裁剪':
            self.plotter[0, 1].add_mesh_clip_box(self.mesh_slice,
                                                 name="model_slice",
                                                 invert=False,  # TODO 是否翻转/反转剪辑的标志
                                                 rotation_enabled=True,  # TODO 如果False，框小部件不能旋转，并且严格正交于笛卡尔轴
                                                 widget_color=None,  # TODO 小工具的颜色。字符串、RGB序列或十六进制颜色字符串，as color='white'
                                                 outline_translation=True,  # TODO 如果False，平面小部件不能被转换，并且被严格地放置在给定的边界内
                                                 merge_points=True,  # TODO 如果True(默认)，独立定义的网格元素的重合点将被合并
                                                 crinkle=False,  # TODO 通过沿剪辑提取整个单元格来使剪辑起皱
                                                 interaction_event='end',
                                                 # TODO 什么时候触发改变，'start'刚开始拖动, 'end'拖动松开, 'always'一直变
                                                 # **kwargs    # TODO 所有add_mesh()的参数都可以用
                                                 cmap="coolwarm",
                                                 show_scalar_bar=False,
                                                 )
            self.plotter[0, 1].add_scalar_bar(title=node_data_name,
                                              color='white',
                                              label_font_size= 14,
                                              vertical=True,
                                              height=0.65,
                                              position_x=0.85,
                                              position_y=0.1
                                              )
        elif sliceType == '数据阈值裁剪':
            self.plotter[0, 1].clear_actors()
            data_name = node_data_name + " "  # TODO 用于设置标量条
            self.plotter[0, 1].add_mesh_threshold(self.mesh_slice,
                                                  name="model_slice2",
                                                  scalars=node_data_name,  # TODO 要设定阈值和显示的网格上标量的字符串名称
                                                  invert=False,  # TODO 反转阈值结果。也就是说，此选项关闭时输出中的像元将被排除，而输出中的像元将被包括
                                                  widget_color=None,  # TODO 小工具的颜色。字符串、RGB序列或十六进制颜色字符串。as color='white'
                                                  preference='cell',
                                                  # TODO 该参数设置标量如何映射到网格。默认'cell'，导致标量与网格单元相关联。可以是'point'或者'cell'
                                                  title=None,  # TODO 滑块小工具的字符串标签
                                                  pointa=(0.4, 0.9),  # TODO 显示端口上滑块左侧点的相对坐标
                                                  pointb=(0.9, 0.9),  # TODO 显示端口上滑块右点的相对坐标
                                                  continuous=False,
                                                  # TODO 如果启用了此选项(默认为False)，使用连续音程[minimum cell scalar, maximum cell scalar]与阈值边界相交，而不是与顶点的离散标量值集合相交
                                                  all_scalars=False,
                                                  # TODO 如果使用点数据的标量，当该值为True时，单元中的所有点必须满足阈值。当False的时候，具有满足阈值标准的标量值的像元的任何点将提取像元。使用单元格数据时无效
                                                  method='upper',
                                                  # TODO 为单值设置阈值方法，定义要使用的阈值界限。如果value是一个范围，此参数将被忽略，提取两个值之间的数据。对于单个值，'lower'将提取低于value. 'upper'将提取大于value.
                                                  # **kwargs    # TODO 所有add_mesh()的参数都可以用
                                                  cmap="coolwarm",
                                                  scalar_bar_args={
                                                      'title': data_name,
                                                      'color':'white',
                                                      'label_font_size': 14,
                                                      "vertical": "False",
                                                      "height": 0.65,
                                                      "position_x": 0.85,
                                                      "position_y": 0.1
                                                  },
                                                  )
        else:
            # print("裁剪类型错误")
            pass

        self.plotter[0, 1].add_axes(  # TODO 左下角坐标轴
            line_width=5,
            color='white',
            cone_radius=0.6,
            shaft_length=0.7,
            tip_length=0.3,
            ambient=0.5,
            label_size=(0.4, 0.16),
        )

        self.plotter[0, 1].add_camera_orientation_widget()  # TODO 右上角坐标轴带负轴

        self.plotter[0, 1].reset_camera()


    # 选行追溯云图 选择框 改变事件
    def click_tableSelectEnable(self):
        if self.ui.checkBox_tableSelectEnable.isChecked():
            # TODO 设置表格 可选中
            self.ui.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        else:
            # TODO 设置表格 不可选中
            self.ui.tableWidget.setSelectionMode(QAbstractItemView.NoSelection)

    # 表格 单元格被选中 事件
    def table_history_selected_item(self):
        global history_vtu_filename, history_vtu_directory
        selected_items = self.ui.tableWidget_History.selectedItems()

        if selected_items:
            selected_row = selected_items[0].row()
            selected_column = selected_items[0].column()
            time_element = self.ui.tableWidget_History.item(selected_row, 0).text() # TODO 表格第1列
            time_element = str(time_element).replace("/", "-").replace(":", "：")
            # print(f"time_element: {time_element}")
            # time_element = int(time_element)
            # time_element = int(time_element) - TIME_TOONE   # TODO 因为归一化，所以 原始数据 减去开始降温的时间 得到路径
            history_vtu_filename = f"{history_vtu_directory}/CylinderModelTime_{time_element}.vtu"
            print(f"选中行的vtu文件名称: {history_vtu_filename}")
            # self.ui.lab_vtuPath.setText(history_vtu_filename)

    # 云图显示类型（表面、网格、线、点）下拉框的 改变事件
    def handleSelectionChange(self):
        selected_text = self.ui.cbox_styleType.currentText()
        print(selected_text)
        if selected_text == 'surface':
            self.ui.cbox_sliceType.clear()
            self.ui.cbox_sliceType.addItems(["盒子裁剪", "数据阈值裁剪"])
        else:
            self.ui.cbox_sliceType.clear()
            self.ui.cbox_sliceType.addItem("无")

    # 保存模型（尺寸、网格）按钮  单击事件
    def click_btn_saveModel(self):
        self.ui.btn_saveModel.setEnabled(False)
        self.ui.btn_resetModel.setEnabled(True)
        mesh = self.modelCustomTool.getModel()
        self.mesh = mesh.copy()
        # 初始化 数据
        zero_list = np.array([0] * len(self.mesh.points))
        self.mesh.point_data["Temperature"] = zero_list
        self.mesh.point_data["r_strain"] = zero_list
        self.mesh.point_data["θ_strain"] = zero_list
        self.mesh.point_data["z_strain"] = zero_list
        self.mesh.point_data["r_θ_z_strain"] = zero_list
        self.mesh.point_data["r_stress"] = zero_list
        self.mesh.point_data["θ_stress"] = zero_list
        self.mesh.point_data["z_stress"] = zero_list
        self.mesh.point_data["r_θ_z_stress"] = zero_list

        self.mesh_slice = self.mesh.copy()
        self.mesh_now = self.mesh.copy()
        # 隐藏滑块
        self.radius_slider.SetEnabled(0)
        self.height_slider.SetEnabled(0)
        self.r_resolution_slider.SetEnabled(0)
        self.theta_resolution_slider.SetEnabled(0)
        self.z_resolution_slider.SetEnabled(0)
        # TODO 空的云图
        self.model_none = self.plotter_now[0,0].add_mesh(mesh, name="model_none",)
        self.plotter_modelModify[0, 0].add_mesh(self.cylinder_mesh, name="modelModify", show_edges=False,)
        # print(f"{self.modelModify_actor}")
        print(f"保存模型{self.mesh}")
        vtuModel.set_model(self.mesh)


    # 重置模型（尺寸、网格）按钮  单击事件
    def click_btn_resetModel(self):
        self.ui.btn_saveModel.setEnabled(True)
        self.ui.btn_resetModel.setEnabled(False)
        # 显示滑块
        self.radius_slider.SetEnabled(1)
        self.height_slider.SetEnabled(1)
        self.r_resolution_slider.SetEnabled(1)
        self.theta_resolution_slider.SetEnabled(1)
        self.z_resolution_slider.SetEnabled(1)
        # self.cylinder_mesh = self.mesh
        # self.radius_slider.GetRepresentation().SetValue(55)
        self.plotter_modelModify[0, 0].add_mesh(self.cylinder_mesh, name="modelModify", show_edges=True,)

    # 选择模型存储文件夹
    def click_openDirectory(self):
        global vtu_storage_directory
        folder_name = QFileDialog.getExistingDirectory(self.ui, "选择模型存储文件夹", "../CylinderData")
        vtu_storage_directory = folder_name
        print(folder_name)

    # 选择监测点配置文件 # TODO 读取 监测点 配置文件，并存储到 view_points_dic 字典中，后续进行分别读取即可
    def click_openFileOfViewPoint(self):
        print(self.ui.btn_saveModel.isEnabled())
        if self.ui.btn_saveModel.isEnabled():
            # 显示警告弹窗
            QMessageBox.warning(self.ui, '警告', '请先保存模型！')
            return
        file_name = QFileDialog.getOpenFileName(self.ui, "选择监测点配置模板文件",
                                                "../CylinderData",
                                                'Excel files (*.xls *.xlsx);; All files (*)')  # 选择文件，返回选中的文件路径
        print(file_name)
        if file_name:
            file_name = file_name[0]
            self.file_name = file_name
            # 根据文件类型读取数据
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                # 读取Excel文件
                try:
                    with pd.ExcelFile(file_name) as file:
                        points_df = pd.read_excel(file, sheet_name="数据列对应")  # 要读取的工作表名
                except:
                    print('模板文件选择错误，请重新选择！')
                    QMessageBox.warning(self.ui, '警告', '模板文件选择错误，请重新选择！')
                    return

            else:
                # 无法识别的文件类型
                print('无法识别的文件类型')
                return
            global view_points_dic

            points_df["theta/mm"] = 0
            # TODO 用于判断theta方向应变 是端面监测点 还是环面监测点
            r_max = points_df["r/mm"].max()
            z_max = points_df["z/mm"].max()
            print(r_max, z_max)
            # TODO 验证监测点模板内容  和  模型尺寸是否一致
            print(f"边界{self.mesh.bounds}")
            # print(self.mesh.bounds[1])  # x最大
            # print(self.mesh.bounds[-1]) # z最大
            if (np.abs(r_max - self.mesh.bounds[1]) >= 1) or (np.abs(z_max - self.mesh.bounds[-1]) >= 1):
                QMessageBox.warning(self.ui, '警告', '监测点位置信息和模型尺寸不匹配！')
                return

            # points_df = points_df[["r/mm", "theta/mm", "z/mm"]]
            # known_points_list = np.array(points_df)
            print(points_df)
            loc_r = points_df.loc[:, "监测类型"] == "r方向应变"
            # print(loc_r)
            top_r_strain_points_df = points_df[loc_r]
            # r_strain_points_df["theta/mm"] = 0
            top_r_strain_points_df = top_r_strain_points_df[
                ["Excel中列名称", "监测类型", "r/mm", "theta/mm", "z/mm"]].sort_values(by='r/mm', ascending=True).reset_index(drop=True)
            # print(top_r_strain_points_df)
            top_r_ViewPoints = ViewPoints(np.array(top_r_strain_points_df["Excel中列名称"]),
                                          top_r_strain_points_df["监测类型"][0],
                                          np.array(top_r_strain_points_df[["r/mm", "theta/mm", "z/mm"]]))
            view_points_dic["r方向应变"] = top_r_ViewPoints # 将监测点对象加入 字典 方便使用

            loc_z = points_df.loc[:, "监测类型"] == "z方向应变"
            # print(loc_z)
            surface_z_strain_points_df = points_df[loc_z]
            # r_strain_points_df["theta/mm"] = 0
            surface_z_strain_points_df = surface_z_strain_points_df[
                ["Excel中列名称", "监测类型", "r/mm", "theta/mm", "z/mm"]].sort_values(by='z/mm', ascending=True).reset_index(drop=True)
            print("=============================================")
            print(surface_z_strain_points_df)
            surface_z_ViewPoints = ViewPoints(np.array(surface_z_strain_points_df["Excel中列名称"]),
                                          surface_z_strain_points_df["监测类型"][0],
                                          np.array(surface_z_strain_points_df[["r/mm", "theta/mm", "z/mm"]]))
            view_points_dic["z方向应变"] = surface_z_ViewPoints # 将监测点对象加入 字典 方便使用

            loc_theta_top = (points_df.loc[:, "监测类型"] == "theta方向应变") & (points_df.loc[:, "z/mm"] == z_max)
            top_theta_strain_points_df = points_df[loc_theta_top]
            top_theta_strain_points_df = top_theta_strain_points_df[
                ["Excel中列名称", "监测类型", "r/mm", "theta/mm", "z/mm"]].sort_values(by='r/mm', ascending=True).reset_index(drop=True)
            print(top_theta_strain_points_df)
            top_theta_ViewPoints = ViewPoints(np.array(top_theta_strain_points_df["Excel中列名称"]),
                                          top_theta_strain_points_df["监测类型"][0],
                                          np.array(top_theta_strain_points_df[["r/mm", "theta/mm", "z/mm"]]))
            view_points_dic["top_theta方向应变"] = top_theta_ViewPoints # 将监测点对象加入 字典 方便使用

            loc_theta_surface = (points_df.loc[:, "监测类型"] == "theta方向应变") & (points_df.loc[:, "r/mm"] == r_max)
            surface_theta_strain_points_df = points_df[loc_theta_surface]
            surface_theta_strain_points_df = surface_theta_strain_points_df[
                ["Excel中列名称", "监测类型", "r/mm", "theta/mm", "z/mm"]].sort_values(by='z/mm', ascending=True).reset_index(drop=True)
            # print(surface_theta_strain_points_df)
            surface_theta_ViewPoints = ViewPoints(np.array(surface_theta_strain_points_df["Excel中列名称"]),
                                          surface_theta_strain_points_df["监测类型"][0],
                                          np.array(surface_theta_strain_points_df[["r/mm", "theta/mm", "z/mm"]]))
            view_points_dic["surface_theta方向应变"] = surface_theta_ViewPoints # 将监测点对象加入 字典 方便使用

            loc_temperature = points_df.loc[:, "监测类型"] == "温度"
            temperature_points_df = points_df[loc_temperature]
            temperature_points_df = temperature_points_df[
                ["Excel中列名称", "监测类型", "r/mm", "theta/mm", "z/mm"]].sort_values(by='r/mm', ascending=True).reset_index(drop=True)
            # print(temperature_points_df)
            temperature_ViewPoints = ViewPoints(np.array(temperature_points_df["Excel中列名称"]),
                                          temperature_points_df["监测类型"][0],
                                          np.array(temperature_points_df[["r/mm", "theta/mm", "z/mm"]]))
            view_points_dic["温度"] = temperature_ViewPoints # 将监测点对象加入 字典 方便使用

            loc_bgtemperature = points_df.loc[:, "监测类型"] == "环境温度"
            bgtemperature_points_df = points_df[loc_bgtemperature]
            bgtemperature_points_df = bgtemperature_points_df[["Excel中列名称", "监测类型"]].reset_index(drop=True)
            # print(bgtemperature_points_df)
            bgtemperature_ViewPoints = ViewPoints(np.array(bgtemperature_points_df["Excel中列名称"]),
                                          bgtemperature_points_df["监测类型"][0],
                                          None)
            view_points_dic["环境温度"] = bgtemperature_ViewPoints # 将监测点对象加入 字典 方便使用

            loc_ae = points_df.loc[:, "监测类型"] == "声发射"
            ae_points_df = points_df[loc_ae]
            ae_points_df = ae_points_df[
                ["Excel中列名称", "监测类型", "r/mm", "theta/mm", "z/mm"]].sort_values(by='r/mm', ascending=True).reset_index(drop=True)
            # print(ae_points_df)
            ae_ViewPoints = ViewPoints(np.array(ae_points_df["Excel中列名称"]),
                                          ae_points_df["监测类型"][0],
                                          np.array(ae_points_df[["r/mm", "theta/mm", "z/mm"]]))
            view_points_dic["声发射"] = ae_ViewPoints # 将监测点对象加入 字典 方便使用
            print(view_points_dic)
            print(view_points_dic["温度"])
            print(view_points_dic["surface_theta方向应变"])
            # TODO 这里需要添加验证view_points_dic中是否有数据，以及r方向应变的Excel列名称列表和top_theta方向应变的Excel列名称列表长度是否一致
            # 验证view_points_dic中是否有数据
            try:
                # r方向应变的Excel列名称列表和top_theta方向应变的Excel列名称列表长度是否一致
                if len(view_points_dic["r方向应变"].column_name_list) != len(
                        view_points_dic["top_theta方向应变"].column_name_list):
                    QMessageBox.warning(self.ui, '警告',
                                        '表格内容填写错误！请检查【r方向应变】和【top_theta方向应变】监测点数量')
                    return
            except:
                QMessageBox.warning(self.ui, '警告',
                                    '监测点Excel模板文件内容有误，请正确填写后再上传！')
                return

            global view_points_mesh_dic     # key:english_column_name，value:mesh
            global view_points_labels_dic     # key:english_column_name，value:标注的坐标
            global view_points_coordinates_dic    # key:english_column_name，value:监测点的坐标
            # 遍历view_points_dic
            for key, view_points in view_points_dic.items():
                print(key, view_points)
                #  获取监测点在模型中的坐标
                coordinates_list = view_points.coordinates_list
                english_name_list = view_points.english_column_name_list
                if coordinates_list is not None and english_name_list is not None:
                    for i in range(len(coordinates_list)):
                        view_mesh = pv.CylinderStructured(
                            radius=np.linspace(0, 1, 50),  # 1半径
                            height=1.0,  # 高度
                            center=(coordinates_list[i][0], 0.0, coordinates_list[i][2]),  # 中心
                            direction=(0.0, 0.0, 1.0) if coordinates_list[i][2] == z_max else (1.0, 0.0, 0.0),  # 高度方向
                            theta_resolution=100,  # θ方向的网格精度
                            z_resolution=30
                        )
                        view_points_mesh_dic[english_name_list[i]] = view_mesh
                        view_points_labels_dic[english_name_list[i]] = [coordinates_list[i][0] + r_max / 2,
                                                                        coordinates_list[i][1] + np.random.randint(-r_max, r_max),
                                                                        coordinates_list[i][2] + z_max / 2]
                        view_points_coordinates_dic[english_name_list[i]] = coordinates_list[i]
                else:
                    continue
            # 遍历view_points_mesh_dic
            for key, view_points_mesh in view_points_mesh_dic.items():
                # print(key, view_points_mesh)
                mesh_color = 'red' if key.startswith('T') else ('blue' if key.startswith('Strain') else 'yellow')
                self.plotter_modelModify[0, 0].add_mesh(view_points_mesh, name=key, label=key, color=mesh_color, )
                self.plotter_now[0, 0].add_mesh(view_points_mesh, name=key, label=key, color=mesh_color, )
            # self.plotter_modelModify[0, 0].add_legend(loc='upper left',
            #                                           size=(0.1, 0.7),
            #                                           bcolor=None,
            #                                           face=None, )
            # 获取view_points_labels_dic的key
            labels_key_list = list(view_points_labels_dic.keys())
            # 获取view_points_labels_dic的value
            labels_coordinates_list = np.array(list(view_points_labels_dic.values()))
            # 获取view_points_coordinates_dic的value
            coordinates_list = np.array(list(view_points_coordinates_dic.values()))
            print('labels_value_list', labels_coordinates_list)
            print('coordinates_list', coordinates_list)

            self.plotter_modelModify[0, 0].add_point_labels(points=labels_coordinates_list,
                                                            labels=labels_key_list,
                                                            name='points_labels',
                                                            italic=True,
                                                            text_color='black',
                                                            # point_color='red',
                                                            show_points=False,
                                                            font_size=22,
                                                            bold=True,
                                                            font_family='times',  # courier times  arial
                                                            shape='rounded_rect',     # rect（矩形）  rounded_rect（圆角矩形）
                                                            shape_color='yellow',   # grey
                                                            shape_opacity=0.3,  # 透明度 0-1
                                                            margin=10,    #  边距
                                                            always_visible=True,    # False
                                                            )
            print(">>>>>>>>>>>>>>>>>>>>")
            # labels_key_list[0] = '1'
            print(len(labels_coordinates_list))
            print(labels_key_list)
            # labels_key_list = ['Strain2:-650.0', 'Strain4:-521.26', 'Strain6:-543.34', 'Strain8:-578.89', 'Strain15:-472.63', 'Strain13:-438.99', 'Strain11:-423.84', 'Strain9:-901.55', 'Strain1:-503.02', 'Strain3:-392.36', 'Strain5:-592.94', 'Strain7:-430.85', 'Strain16:-647.85', 'Strain14:-508.94', 'Strain12:-458.14', 'Strain10:-826.22', 'T1:51.18', 'T2:51.28', 'T3:51.3', 'T4:51.45', 'T5:51.74', 'T6:51.48', 'T7:51.55', 'T8:51.26', 'AE1:0', 'AE2:0', 'AE3:0']
            global point_label_actor
            point_label_actor = self.plotter_now[0, 0].add_point_labels(points=labels_coordinates_list,
                                                            labels=labels_key_list,
                                                            name='points_labels',
                                                            italic=True,
                                                            text_color='black',
                                                            # point_color='red',
                                                            show_points=False,
                                                            font_size=22,
                                                            bold=True,
                                                            font_family='times',  # courier times  arial
                                                            shape='rounded_rect',  # rect（矩形）  rounded_rect（圆角矩形）
                                                            shape_color='yellow',  # grey
                                                            shape_opacity=0.3,  # 透明度 0-1
                                                            margin=10,  # 边距
                                                            always_visible=True,
                                                            )
            # 将监测点坐标和标注坐标组合
            view_points_lines_list = []
            for i in range(len(coordinates_list)):
                view_points_lines_list.append(coordinates_list[i])
                view_points_lines_list.append(labels_coordinates_list[i])
            view_points_lines_list = np.array(view_points_lines_list)
            # print(view_points_lines_list)
            # 添加标注 线
            self.plotter_modelModify[0, 0].add_lines(view_points_lines_list, name='point_label_lines', color='purple', width=3)
            self.plotter_now[0, 0].add_lines(view_points_lines_list, name='point_label_lines', color='purple',
                                                     width=3)
            # TODO 当前时刻的云图
            self.mesh_now.point_data["Temperature"] = np.array([0] * len(self.mesh_now.points))
            self.mesh_now.set_active_scalars("Temperature")
            self.model_now = self.plotter_now[0, 0].add_mesh(self.mesh_now, name="model_now", cmap="coolwarm",
                                                             show_scalar_bar=False,
                                                             # show_vertices=True,  # TODO   显示点
                                                             point_size=5,  # TODO   点的大小
                                                             style='surface',
                                                             # TODO  显示样式默认：'surface'表面，'wireframe'线条，'points'表面的点配合point_size设置大小，
                                                             )  # TODO   圆柱
            self.model_now.visibility = False

    # 选择实验数据文件
    def click_openFile(self):
        global vtu_storage_directory
        print(vtu_storage_directory)
        if vtu_storage_directory is None or vtu_storage_directory=='':
            # 显示警告弹窗
            QMessageBox.warning(self.ui, '警告', '请先选择模型存储文件夹！')
            return

        file_name = QFileDialog.getOpenFileName(self.ui, "选择实验数据文件", "../CylinderData",
                                               'Excel files (*.xls *.xlsx);; All files (*)')  # 选择文件，返回选中的文件路径
        print(file_name)
        if file_name:
            file_name = file_name[0]
            self.file_name = file_name
            # 根据文件类型读取数据
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                # 读取Excel文件
                # data = pd.read_excel(file_name)
                # self.set_table_data(data)
                global target_file_path

                target_file_path = file_name
                # 开始后，设置按钮不可用
                self.ui.btn_saveModel.setEnabled(False)     # 保存模型按钮
                self.ui.btn_resetModel.setEnabled(False)    # 重置模型按钮
                self.ui.viewPointButton.setEnabled(False)   # 导入监测点配置 按钮
                self.ui.selectFDirectoryButton.setEnabled(False)    # 选择模型存储文件夹 按钮
                self.ui.selectFileButton.setEnabled(False)      # 选择实验数据文件 按钮
                # t_write.start()
                t_read.start()
                t_write_vtu.start()
                t_plot.start()
            else:
                # 无法识别的文件类型
                print('无法识别的文件类型')
                return



    # 选择历史云图存放文件夹 按钮 单击事件
    def click_selectHistoryVtuDirectory(self):
        isSaved = self.ui.btn_saveModel.isEnabled()  # 保存模型按钮
        print(f"isSaved----{isSaved}")
        if isSaved:
            # 显示警告弹窗
            QMessageBox.warning(self.ui, '警告', '请先保存模型！')
            return

        global history_vtu_directory
        # 获取 历史云图存放文件夹 路径
        history_vtu_directory = QFileDialog.getExistingDirectory(self.ui, "选择历史云图存放文件夹", "../CylinderData")
        print(history_vtu_directory)
        if history_vtu_directory is not None or history_vtu_directory != '':
            # 清空tableWidget_History表格
            self.ui.tableWidget_History.clear()

    # 选择历史实验数据文件
    def click_selectHistoryDataFile(self):
        global history_vtu_directory
        print(history_vtu_directory)
        if history_vtu_directory is None or history_vtu_directory == '':
            # 显示警告弹窗
            QMessageBox.warning(self.ui, '警告', '请先选择历史云图存放文件夹！')
            return

        file_name = QFileDialog.getOpenFileName(self.ui, "选择历史实验数据文件", "../CylinderData",
                                                'Excel files (*.xls *.xlsx);; All files (*)')  # 选择文件，返回选中的文件路径
        print(file_name)
        if file_name:
            file_name = file_name[0]
            self.file_name = file_name
            # 根据文件类型读取数据
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                # 清空tableWidget_History表格
                self.ui.tableWidget_History.clear()
                # 读取Excel文件
                data = pd.read_excel(file_name)
                # print(data)
            else:
                # 无法识别的文件类型
                print('无法识别的文件类型')
                return

            self.set_history_table_data(data)

    # 给历史数据表格添加数据
    def set_history_table_data(self, data):  # data:DataFrame
        global history_vtu_directory
        enable_vtu_list = get_enable_vtu_list(history_vtu_directory)  # 可用的vtu列表，['2023-10-24 09：22：12', ]
        # print(f"enable_vtu_list: {enable_vtu_list}")
        enable_vtu_list = [vtu_time.replace("-", "/").replace("：", ":") for vtu_time in enable_vtu_list]
        if data.shape[0] != 0:  # 如果有数据才添加
            # 使用 apply 函数检查整个DataFrame每行的# TODO 第2个元素 是否在列表中，并将布尔值转换为字符串
            data = data[data["采样时间"].dt.strftime('%Y/%m/%d %H:%M:%S').isin(enable_vtu_list)]  # 筛选出'采样时间'列的值在enable_vtu_list中的行

        # 将数据添加到表格中
        self.ui.tableWidget_History.setRowCount(data.shape[0])
        self.ui.tableWidget_History.setColumnCount(data.shape[1])
        # 设置水平表头
        header = data.columns.tolist()

        self.ui.tableWidget_History.setHorizontalHeaderLabels(header)
        for i in range(data.shape[0]):
            for j in range(data.shape[1]):
                item = QTableWidgetItem(str(data.iloc[i, j]))
                item.setTextAlignment(Qt.AlignCenter)
                # 设置单元格不可编辑
                # item.setFlags(Qt.ItemIsEnabled)
                self.ui.tableWidget_History.setItem(i, j, item)
        # TODO 优化 5 将行与列的高度设置为所显示的内容的宽度高度匹配
        QTableWidget.resizeColumnsToContents(self.ui.tableWidget_History)
        QTableWidget.resizeRowsToContents(self.ui.tableWidget_History)

    # 给表格添加数据
    def set_table_data(self, data): # data:DataFrame
        # 将数据添加到表格中
        self.ui.tableWidget.setRowCount(data.shape[0])
        self.ui.tableWidget.setColumnCount(data.shape[1])
        # 设置水平表头
        header = data.columns.tolist()
        self.ui.tableWidget.setHorizontalHeaderLabels(header)
        for i in range(data.shape[0]):
            for j in range(data.shape[1]):
                item = QTableWidgetItem(str(data.iloc[i, j]))
                item.setTextAlignment(Qt.AlignCenter)
                # 设置单元格不可编辑
                # item.setFlags(Qt.ItemIsEnabled)
                self.ui.tableWidget.setItem(i, j, item)
        if self.ui.checkBox_tableAutoScroll.isChecked():    # 表格自动滚动选择框 被选中
            # 通过鼠标滚轮定位，快速定位到最后一行
            self.ui.tableWidget.verticalScrollBar().setSliderPosition(self.ui.tableWidget.rowCount() - 1)
        # TODO 优化 5 将行与列的高度设置为所显示的内容的宽度高度匹配
        QTableWidget.resizeColumnsToContents(self.ui.tableWidget)
        QTableWidget.resizeRowsToContents(self.ui.tableWidget)

    # 更新 当前时刻的云图  mesh是固定的，每次读取文件 重新给节点数据赋值
    def update_plotter_now(self, now_vtu_filename):
        # 读取 VTK 文件
        mesh = pv.read(now_vtu_filename)
        # 读取节点数据
        node_data_name = "Temperature"  # 替换为您的节点数据名称  Temperature
        node_data_name = self.ui.cbox_predict_dataType.currentText()    # 获取选择的数据类型
        node_data = mesh.point_data[node_data_name]
        # print(f"node_data_name=======>>>>>>>>{node_data_name}")
        # print(f"node_data=======>>>>>>>>{node_data}")
        # print(self.model_now)
        self.mesh_now.set_active_scalars(node_data_name)
        self.mesh_now.point_data[node_data_name] = node_data
        self.model_none.visibility = False  # 隐藏空的云图
        self.model_now.visibility = True    # 显示当前的云图
        self.plotter_now[0,0].add_scalar_bar(
                                    # title= node_data_name,  #"r_strain"
                                    color='white',
                                    label_font_size=14,
                                    vertical= True,
                                    height= 0.65,
                                    position_x= 0.1,   # 0.85
                                    position_y= 0.2
                                    )
        licm = (np.min(node_data), np.max(node_data))
        # print(licm)
        self.plotter_now[0,0].update_scalar_bar_range(licm)  # , name="Temperature"

    #  更新当前云图的监测点显示数据
    def update_plotter_now_labels(self, labels_coordinates_list, now_labels_list):
        global point_label_actor
        # print(f"第{i}次{len(labels_coordinates_list)}")
        # print(f"第{i}次{labels_coordinates_list}")
        # print(f"第{i}次{now_labels_list}")
        # self.plotter_now[0, 0].remove_actor(point_label_actor) if point_label_actor is not None else None
        point_label_actor = self.plotter_now[0, 0].add_point_labels(points=labels_coordinates_list,
                                                labels=now_labels_list,
                                                name='points_labels',
                                                italic=True,
                                                text_color='black',
                                                # point_color='red',
                                                show_points=False,
                                                font_size=22,
                                                bold=True,
                                                font_family='times',  # courier times  arial
                                                shape='rounded_rect',  # rect（矩形）  rounded_rect（圆角矩形）
                                                shape_color='yellow',  # grey
                                                shape_opacity=0.3,  # 透明度 0-1
                                                margin=10,  # 边距
                                                always_visible=True,
                                                )

    #  更新 云图选中点 显示数据
    def update_pickPoint_label(self, pickPoint, pickPoint_index):
        node_data_name = self.ui.cbox_predict_dataType.currentText()  # 获取选择的数据类型
        node_data = self.mesh_now.point_data[node_data_name]
        # print(f"update_pickPoint_label[node_data_name]{node_data_name}")
        # print(f"update_pickPoint_label[node_data]{node_data}")
        # mesh = pv.Cube(center=point, x_length=1, y_length=1, z_length=1)
        # window.plotter_now[0,0].add_mesh(mesh, name='cloud_now_picked', style='wireframe', color='r')
        self.plotter_now[0, 0].add_point_labels(pickPoint, [f"{node_data[pickPoint_index]:.2f}"], name='cloud_now_picked',
                                                  italic=True,
                                                  text_color='black',
                                                  # point_color='red',
                                                  show_points=False,
                                                  font_size=22,
                                                  bold=True,
                                                  font_family='times',  # courier times  arial
                                                  shape='rounded_rect',  # rect（矩形）  rounded_rect（圆角矩形）
                                                  shape_color='yellow',  # grey
                                                  shape_opacity=0.3,  # 透明度 0-1
                                                  margin=10,  # 边距
                                                  always_visible=True,
                                                  )


    # 获取拟合参数里的 弃用
    def flash_plotter_now(self, fitting_data_list):
        self.mesh_now.set_active_scalars("Temperature")
        self.mesh_now.point_data["Temperature"] = fitting_data_list

        self.model_none.visibility = False  # 隐藏空的云图
        self.model_now.visibility = True  # 显示当前的云图
        self.plotter_now[0, 0].add_scalar_bar(
            # title= node_data_name,  #"r_strain"
            color='white',
            label_font_size=14,
            vertical=True,
            height=0.65,
            position_x=0.85,
            position_y=0.1
        )
        licm = (np.min(fitting_data_list), np.max(fitting_data_list))
        # print(licm)
        self.plotter_now[0, 0].update_scalar_bar_range(licm)  # , name="Temperature"


    # 设置 axis的 x，y坐标范围
    def set_axis_xylim(self, ax, x_data_list, view_points):
        # 获取 x ，y轴刻度位置
        xticks = ax.get_xticks()
        yticks = ax.get_yticks()
        # 计算相邻刻度位置之间的差异
        xtick_size = mdate.num2date(xticks[1]) - mdate.num2date(xticks[0])
        ytick_size = yticks[1] - yticks[0]
        xtick_first_date = datetime.strptime(x_data_list[0], "%Y/%m/%d %H:%M:%S")
        predic_lenth = len(view_points.predict_time)
        last_date = datetime.strptime(x_data_list[-1], "%Y/%m/%d  %H:%M:%S")  # 当前时间（预测开始时间）
        xtick_last_date = last_date + timedelta(seconds=predic_lenth)
        ax.set_xlim(xtick_first_date - xtick_size,
                                      xtick_last_date + xtick_size)

        ax.set_ylim(
            np.min([np.min(view_points.data), np.min(view_points.predict_data)]) - ytick_size,
            np.max([np.max(view_points.data), np.max(view_points.predict_data)]) + ytick_size)

    # 更新 温度 曲线图 type: "温度" or "应变"
    def update_temperature_strain_charts(self, x_data_list, view_points, type):
        global mouse_in_axis, colors    # 鼠标是否在坐标轴内
        if view_points is not None:
            shape = view_points.data.shape[1]
            half_shape = int(shape / 2)
            # 颜色列表
            previous_colors = []
            for i in range(shape+3):    #  +3 是为了防止 颜色列表 不够用
                # 调用函数  生成一个带有足够颜色通道差异的随机RGB颜色，确保与之前生成的颜色不相邻。
                color = generate_random_color(min_distance=0.3, previous_colors=previous_colors)
                if color is not None:
                    previous_colors.append(color)
            # print(f"previous_colors===>>{previous_colors}")
            color_i = 0
            for i in range(shape):
                name = view_points.column_name_list[i]
                if i < half_shape:
                    # 判断是温度 还是应变
                    if type == "温度":
                        # 如果鼠标不在子图内，重新设置坐标轴刻度
                        self.set_axis_xylim(self.ax_temperature1, x_data_list,
                                            view_points) if not mouse_in_axis else None

                        # Matplotlib中的
                        self.create_update_charts_line(name=name, ax=self.ax_temperature1,
                                                       lines=self.ax_temperature1_lines,
                                                       local_color=previous_colors[i], x_data_list=x_data_list,
                                                       view_points=view_points, i=i)
                    else:   # 应变
                        # 如果鼠标不在子图内，重新设置坐标轴刻度
                        self.set_axis_xylim(self.ax_strain1, x_data_list,
                                            view_points) if not mouse_in_axis else None

                        # Matplotlib中的
                        self.create_update_charts_line(name=name, ax=self.ax_strain1,
                                                       lines=self.ax_strain1_lines,
                                                       local_color=previous_colors[i], x_data_list=x_data_list,
                                                       view_points=view_points, i=i)



                else:
                    # 判断是温度 还是应变
                    if type == "温度":
                        # 如果鼠标不在子图内，重新设置坐标轴刻度
                        self.set_axis_xylim(self.ax_temperature2, x_data_list,
                                            view_points) if not mouse_in_axis else None

                        # Matplotlib中的
                        self.create_update_charts_line(name=name, ax=self.ax_temperature2,
                                                       lines=self.ax_temperature2_lines,
                                                       local_color=previous_colors[i], x_data_list=x_data_list,
                                                       view_points=view_points, i=i)
                    else:  # 应变
                        # 如果鼠标不在子图内，重新设置坐标轴刻度
                        self.set_axis_xylim(self.ax_strain2, x_data_list,
                                            view_points) if not mouse_in_axis else None

                        # Matplotlib中的
                        self.create_update_charts_line(name=name, ax=self.ax_strain2,
                                                       lines=self.ax_strain2_lines,
                                                       local_color=previous_colors[i], x_data_list=x_data_list,
                                                       view_points=view_points, i=i)
        else:
            print("view_points is None")

    # Pyvista中的
    def update_pyvista_charts(self, x_list, T_index, r_index, theta_index, z_index):
        # 获取 监测点对象
        T_view_point = view_points_dic.get("温度")
        r_strain_view_point = view_points_dic.get("r方向应变")
        theta_strain_view_point = view_points_dic.get("top_theta方向应变")
        z_strain_view_point = view_points_dic.get("z方向应变")
        # 更新温度曲线图
        self.update_pyvista_charts_line(self.temperature_history_line, self.temperature_future_line,
                                        x_list[:], T_view_point, T_index)
        # 更新r应变曲线图
        self.update_pyvista_charts_line(self.r_strain_history_line, self.r_strain_future_line,
                                        x_list[:], r_strain_view_point, r_index)
        # 更新theta应变曲线图
        self.update_pyvista_charts_line(self.theta_strain_history_line, self.theta_strain_future_line,
                                        x_list[:], theta_strain_view_point, theta_index)
        # 更新z应变曲线图
        self.update_pyvista_charts_line(self.z_strain_history_line, self.z_strain_future_line,
                                        x_list[:], z_strain_view_point, z_index)


    # Pyvista中的
    def update_pyvista_charts_line(self, history_line, future_line, x_data_list, view_points, column_index):
        #  获取监测点数据
        data_list = view_points.data[:, column_index]
        pre_data_list = view_points.predict_data[:, column_index]
        pre_time_list = view_points.predict_time[:]
        # 防止时间列表和数据列表长度不一致
        if len(x_data_list) > len(data_list):
            x_data_list = x_data_list[:len(data_list)]
        elif len(x_data_list) < len(data_list):
            data_list = data_list[:len(x_data_list)]

        # 防止时间列表和数据列表长度不一致
        if len(pre_time_list) > len(pre_data_list):
            pre_time_list = pre_time_list[:len(pre_data_list)]
        elif len(pre_time_list) < len(pre_data_list):
            pre_data_list = pre_data_list[:len(pre_time_list)]
        history_line.update(x_data_list[:], data_list[:])
        future_line.update(pre_time_list[:], pre_data_list[:])

    # Matplotlib中的
    def create_update_charts_line(self, ax, lines, local_color, name, x_data_list, view_points, i):
        # TODO 判断是否有线，如果没有实例化一个，有就直接用，如果不判断，会一直刷新绘图
        if lines.get(f"{name}") is None:
            lines[f"{name}"], = ax.plot([0], [0],
                                          color=local_color,
                                          label=name.split("(")[0] if len(name.split("(")) > 0 else name,
                                          ls='-')  # 初始为空数据
            lines[f"{name}_predict"], = ax.plot([0], [0],
                                                  color=local_color,
                                                  # label='预测' + name.split("(")[0] if len(name.split("(")) > 0 else name,
                                                  ls="--")
        # numpy.vectorize用于将普通函数转换为可用于 Numpy 数组的向量化函数。向量化函数能够逐元素操作 Numpy 数组，而不需要显式地使用循环。
        datetime_objects = np.vectorize(lambda time: datetime.strptime(time, "%Y/%m/%d %H:%M:%S"))(x_data_list)
        data_i = view_points.data[:, i]
        predict_data_i = view_points.predict_data[:, i]
        # 防止时间列表和数据列表长度不一致
        if len(datetime_objects) > len(data_i):
            datetime_objects = datetime_objects[:len(data_i)]
        elif len(datetime_objects) < len(data_i):
            data_i = data_i[:len(datetime_objects)]

        one_second = timedelta(seconds=1)
        predic_lenth = len(view_points.predict_time)
        last_date = datetime.strptime(x_data_list[-1], "%Y/%m/%d  %H:%M:%S")  # 当前时间（预测开始时间）
        predic_dates = mdate.drange(last_date, last_date + timedelta(seconds=predic_lenth - 1), one_second)
        # 防止时间列表和数据列表长度不一致
        if len(predic_dates) > len(predict_data_i):
            predic_dates = predic_dates[:len(predict_data_i)]
        elif len(predic_dates) < len(predict_data_i):
            predict_data_i = predict_data_i[:len(predic_dates)]

        lines[f"{name}"].set_xdata(datetime_objects)
        lines[f"{name}"].set_ydata(data_i)

        lines[f"{name}_predict"].set_xdata(predic_dates)
        lines[f"{name}_predict"].set_ydata(predict_data_i)



    # 测试按钮单击事件
    def click_start(self):
        global t_text
        t_text.start()
        # labels_coordinates_list = np.array(list(view_points_labels_dic.values()))
        # now_labels_list = [str((i+1) * np.random.randint(1, 20)) for i in range(len(labels_coordinates_list))]
        # window.update_plotter_now_labels(labels_coordinates_list, now_labels_list)


    # 生成一个带有足够颜色通道差异的随机RGB颜色，确保与之前生成的颜色不相邻。
def generate_random_color(min_distance=0.2, previous_colors=None, max_attempts=100):
    """
    生成一个带有足够颜色通道差异的随机RGB颜色，确保与之前生成的颜色不相邻。

    参数:
        min_distance (float): 颜色通道值之间的最小差异。默认为0.2。
        previous_colors (list): 之前生成的颜色列表。默认为None。
        max_attempts (int): 尝试生成颜色的最大次数。默认为100。

    返回:
        tuple: 表示RGB颜色的元组 (r, g, b)，其中 r、g、b 是在 [0, 1] 范围内的浮点数。
    """

    def color_distance(color1, color2):
        """计算两个颜色之间的欧氏距离。"""
        return math.sqrt(sum((c1 - c2) ** 2 for c1, c2 in zip(color1, color2)))

    if previous_colors is None:
        previous_colors = []

    if not previous_colors:
        # 如果 previous_colors 为空，直接生成一个随机颜色并返回
        return (random.random(), random.random(), random.random())

    for _ in range(max_attempts):
        # 随机生成颜色通道值
        r = random.random()
        g = random.random()
        b = random.random()

        # 检查是否接近黑色(0,0,0)
        if all(channel < 0.1 for channel in (r, g, b)):
            continue  # 如果接近黑色，则重新生成

        # 计算与之前生成的颜色的最小距离
        min_distance_to_previous = min(color_distance(new_color, prev_color) for new_color in ((r, g, b),) for prev_color in previous_colors)

        # 如果足够不相邻，返回颜色
        if min_distance_to_previous >= min_distance:
            return (r, g, b)

    # 如果达到最大尝试次数仍无法生成足够不相邻的颜色，返回 None 或者触发错误，根据需要进行处理。
    return None



# 获取指定文件夹内的vtu文件名称列表 对应的时刻 列表['2023-10-24 09：22：12', ],用于给表格增加一列，用于选中 云图追溯
def get_enable_vtu_list(folder_path):
    # 获取指定文件夹内后缀名为'.vtu'的文件列表
    vtu_files = [file for file in os.listdir(folder_path) if file.endswith('.vtu')]
    enable_show_list = []
    for vtu_file in vtu_files:
        # 去掉后缀名，然后用'_'分割文件名称
        file_name, _ = os.path.splitext(vtu_file)
        if 'CloudFittingError' in file_name:
            continue
        file_parts = file_name.split('_')
        # print(file_parts)
        # 判断文件名称第2项是否为整数
        # if len(file_parts) == 2 and file_parts[1].isdigit():
        if len(file_parts) == 2:
            enable_show_list.append(str(file_parts[1]))
            # 打印符合条件的文件名称
            # print(f"File: {vtu_file}, Second Part: {file_parts[1]}")
    return enable_show_list


def thread_text():
    for i in range(10):
        labels_coordinates_list = np.array(list(view_points_labels_dic.values()))
        now_labels_list = [str((i + 1) * np.random.randint(1, 20)) for i in range(len(labels_coordinates_list))]

        # window.update_plotter_now_labels(labels_coordinates_list, now_labels_list, i)
        # 此处不再直接操作主界面，而是发射信号给MySignals函数中的text_print，
        # 传递参数包括要操作哪个位置和要操作的内容
        window.ms.text_print.emit(labels_coordinates_list, now_labels_list)
        time.sleep(1)


    '''
    vtu_storage_directory2 = '../CylinderData/Text1'
    filename_list = []
    for end_time in range(246, 265):
        now_vtu_filename = f"{vtu_storage_directory2}/CylinderModelTime-{end_time}.vtu"
        filename_list.append(now_vtu_filename)

        window.update_plotter_now(now_vtu_filename)
        
        # 读取 VTK 文件
        mesh = pv.read(now_vtu_filename)
        # 读取节点数据
        node_data_name = "Temperature"  # 替换为您的节点数据名称
        node_data = mesh.point_data[node_data_name]

        plotter.add_mesh(mesh, name="model_now", scalars=node_data, cmap="coolwarm",
                         scalar_bar_args={
                             'title': "Temperature",
                             "vertical": "False",
                             "height": 0.65,
                             "position_x": 0.85,
                             "position_y": 0.1},
                         )
        
        time.sleep(1)
    '''
# 根据监测点模板Excel数据，将实验数据中对应的数据添加到字典中
def extract_data(key, view_points_dic, data_df):
   if view_points_dic.get(key) is not None:
       column_names = view_points_dic.get(key).column_name_list
       data = None
       try:
           data = np.array(data_df[column_names])
       except KeyError:
           print("没有找到列，监测点Excel数据列名和实验数据列名不匹配")
       view_points_dic.get(key).data = data
       return data
   else:
       print("没有找到监测点，请检查监测点Excel数据")
       return None

# 应变归一化
def normalize_strain(direction_strain_lists, direction_name, view_points_dic):
    direction_to_one = direction_strain_lists[0, :]
    # direction_strain_lists = direction_strain_lists - direction_to_one
    view_points_key = f"{direction_name}方向应变"

    view_points_dic.get(view_points_key).data = direction_strain_lists
    view_points_dic.get(view_points_key).toOne_value = direction_to_one


# 读取线程 1秒读取一行存入全局变量中，并显示到界面的表格中
def read_row_by_second():
    event.clear() # 初始设为红灯  相当于flag=False，阻塞绘图线程
    is_pred = 0 # 只判断==1时，用于将图形提示置空
    while True:
        # lock_global.acquire()   # 上锁
        # print("read...上锁")

        global data_excel, bg_Temperature_list, time_list, strain_lists, start_time_of_decrease, end_time, target_file_path,\
                test_Temperature_lists, TIME_TOONE, STRAIN_LIST_TOONE, view_points_dic, decrease_index, real_time_list

        event_break.wait()  # TODO 断裂后，阻塞线程

        lock_file.acquire()  # 锁上
        print("我正在读取，上锁")
        current_index = len(data_excel)
        print(f"current_index={current_index}")
        # 使用 with 语句打开 Excel 文件
        print(window.file_name)
        try:
            # 使用上下文管理器读取 Excel 文件
            with pd.ExcelFile(target_file_path) as file:
                data_df = pd.read_excel(file, sheet_name="Sheet1")  # 请替换成你的工作表名
        except BaseException as e:
            print("读取失败,文件占用, 开锁")
            lock_file.release()  # 开锁
            continue
        # data_df = pd.read_excel(file_name)
        data_excel = np.array(data_df.loc[:, :])
        # return np.array(my_data_df)
        lenth = len(data_excel)
        if lenth > 0:
            print(f"readLenth={lenth}")
            print(f"Times={data_excel[lenth - 1, 0]}")
            window.set_table_data(data_df)
        else:
            lock_file.release()  # 开锁
            print("我读取完毕，开锁")
            time.sleep(1)
            continue
        lock_file.release()  # 开锁
        print("我读取完毕，开锁")


        # bg_Temperature_list = data_excel[:, -1]  # TODO 环境温度，默认最后一列，后期再修改
        # test_Temperature_lists = data_excel[:, 17:25]  # TODO 检测温度，默认17-24，后期再修改
        # time_list = data_excel[:, 1]  # TODO 时间，默认第二列，后期再修改
        # strain_lists = data_excel[:, 2:18]  # TODO 应变，默认第3~18列，后期再修改

        # TODO 根据监测点信息（模板导入）提取数据列表

        real_time_list = np.array(data_df["采样时间"].dt.strftime('%Y/%m/%d %H:%M:%S'))     # '%Y-%m-%d %H:%M:%S'
        # real_time_list = np.array(data_df["采样时间"])     # '%Y-%m-%d %H:%M:%S'
        # real_time_list = np.datetime_as_string(real_time_list, unit='s')  # 's' 表示秒为单位
        # print(f"real_time_list>>>>>{real_time_list}")
        # print(type(real_time_list[0])) if len(real_time_list) > 0 else None
        # for key, value in view_points_dic.items():
        #     print(f"Key: {key}")
        #     print(value)
        #     print("-" * 30)

        if start_time_of_decrease == 0 and decrease_index == 0:     # 判断之前有没有开始降温
            # print("还未开始降温!")
            text_pt.set_text("还未开始降温!")
            filtered_columns = data_df.filter(regex=re.compile(r'^Time'))
            if filtered_columns is not None:
                time_list = np.array(filtered_columns.iloc[:, 0])
            bg_Temperature_list = extract_data("环境温度", view_points_dic, data_df)
            # 更新图形
            # fig.canvas.draw()
            try:
                fig.canvas.draw()
            except BaseException as e:
                print(f"还未开始降温 fig.canvas.draw() error：{e}")
            start_time_of_decrease, decrease_index = get_start_time(time_list)  # 加入当前时刻，计算开始降温的时刻
            # print(f"start_time_of_decrease=====>>>>{start_time_of_decrease}")
            # print(f"time_list======>>>>>{time_list}")
            # start_time_of_decrease = time_list[start_time_of_decrease]
        else:
            decrease_data_df = data_df.iloc[decrease_index:, :] # 截取从降温开始的数据
            print("开始降温的时间：{}".format(real_time_list[decrease_index]))
            filtered_columns = decrease_data_df.filter(regex=re.compile(r'^Time'))
            if filtered_columns is not None:
                time_list = np.array(filtered_columns.iloc[:, 0])
            bg_Temperature_list = extract_data("环境温度", view_points_dic, decrease_data_df)
            test_Temperature_lists = extract_data("温度", view_points_dic, decrease_data_df)
            r_strain_lists = extract_data("r方向应变", view_points_dic, decrease_data_df)
            top_theta_lists = extract_data("top_theta方向应变", view_points_dic, decrease_data_df)
            surface_theta_lists = extract_data("surface_theta方向应变", view_points_dic, decrease_data_df)
            z_strain_lists = extract_data("z方向应变", view_points_dic, decrease_data_df)
            AE_lists = extract_data("声发射", view_points_dic, decrease_data_df)
            # TODO 检查声发射信号，判断是否断裂
            # 定义一个函数，将字符串转换为数字，非数字返回 None
            def to_numeric(value):
                try:
                    return float(value)
                except ValueError:
                    return None

            # TODO 断裂判断
            # 使用函数转换数组中的元素
            numeric_array = np.vectorize(to_numeric)(AE_lists)
            # 获取声发射信号大于30的行的索引
            row_indices, col_indices = np.where(numeric_array > 30)
            break_index = None  # 最早断裂下标
            if len(row_indices) > 0:
                break_index = row_indices[0]
                # 计算断裂绝对时间
                real_time_break = real_time_list[decrease_index + break_index]
                # print(f"real_time_break: {real_time_break}")
                # 计算环境温降
                bg_T_decrease = bg_Temperature_list[0] - bg_Temperature_list[break_index]
                # print(f"bg_T_decrease: {bg_T_decrease}")
                # 计算试样测点（监测点）温降
                T_decrease_list = test_Temperature_lists[0] - test_Temperature_lists[break_index]
                # print(f"T_decrease_list: {T_decrease_list}")
                # 提取监测点 应变/应力
                r_strain_break_list = r_strain_lists[break_index]
                top_strain_break_list = top_theta_lists[break_index]
                surface_strain_break_list = surface_theta_lists[break_index]
                z_strain_break_list = z_strain_lists[break_index]
                # print(f"r_strain_break_list: {r_strain_break_list}")
                # print(f"top_strain_break_list: {top_strain_break_list}")
                # print(f"surface_strain_break_list: {surface_strain_break_list}")
                # print(f"z_strain_break_list: {z_strain_break_list}")
                event_break.clear()  # 设为红灯(False)  相当于flag=False，阻塞  读取线程、预测线程、写入vtu线程
                E = 1.2e+10
                formatted_T_decrease_list = [f'{temperature:.2f}' for temperature in T_decrease_list]
                formatted_r_strain_break_list = [f'{r_strain_break:.2f}' for r_strain_break in r_strain_break_list]
                formatted_r_stress_break_list = [f'{r_strain_break * E:.2f}' for r_strain_break in r_strain_break_list]
                QMessageBox.warning(window.ui, '发生断裂！', f'<font size="4">环境温降：{bg_T_decrease[0]:.2f}</font><br>'
                                                            f'<font size="4">试样测点温降：{", ".join(formatted_T_decrease_list)}</font><br>'
                                                            f'<font size="4">危险点应变：{", ".join(formatted_r_strain_break_list)}</font><br>'
                                                            f'<font size="4">危险点应力：{", ".join(formatted_r_stress_break_list)}</font>')

            # if TIME_TOONE == 0 and STRAIN_LIST_TOONE == None:
            #     pass
            # TODO 时间归一化处理
            TIME_TOONE = start_time_of_decrease
            time_list = time_list - TIME_TOONE
            # print(f"timelist==={time_list}")
            # TODO 应变归一化处理
            r_strain_toOne = r_strain_lists[0, :]
            STRAIN_LIST_TOONE = r_strain_toOne  # 用于判断应变是否已经归一化

            # 归一化 r方向应变
            normalize_strain(r_strain_lists, "r", view_points_dic)
            # 归一化 top_theta方向应变
            normalize_strain(top_theta_lists, "top_theta", view_points_dic)
            # 归一化 surface_theta方向应变
            normalize_strain(surface_theta_lists, "surface_theta", view_points_dic)
            # 归一化 z方向应变
            normalize_strain(z_strain_lists, "z", view_points_dic)

            # TODO 将开始降温时刻

            # 截取几分钟数据用于进行模型训练（数据拟合）
            # number_minutes = 7
            # end_time = start_time_of_decrease + 60 * number_minutes
            end_time = int(time_list[-1])
            # time_second = (end_time - start_time_of_decrease + 1)
            time_second = end_time  # TODO 因为归一化了，所以当前时刻就是已经降温的时间
            area_time = 1
            # if time_second / 60 > area_time:   # TODO 降温area_time分钟后开始 预测并绘图 每秒存入vtu文件
            if time_second > 60:
                event.set() # 设为绿灯，放行绘图线程
                print(f"降温已经：{time_second}秒了")
                is_pred += 1    # 只判断==1时，用于将图形提示置空
                if is_pred == 1:
                    text_pt.set_text("")
                    # 更新图形
                    # fig.canvas.draw()
                    try:
                        fig.canvas.draw()
                    except BaseException as e:
                        print(f"降温已经xxxs fig.canvas.draw() error：{e}")
            else:
                text_pt.set_text(f"开始降温不足{area_time}分钟\n【{area_time*60 -time_second}s】后开始预测")
                # 更新图形
                # fig.canvas.draw()
                try:
                    fig.canvas.draw()
                except BaseException as e:
                    print(f"开始降温不足几分钟 fig.canvas.draw() error：{e}")
        global view_points_mesh_dic  # key:english_column_name，value:mesh
        global view_points_labels_dic  # key:english_column_name，value:标注的坐标
        global view_points_coordinates_dic  # key:english_column_name，value:监测点的坐标
        labels_coordinates_list = np.array(list(view_points_labels_dic.values()))       # 监测点的标注的坐标
        now_labels_list = np.array([])

        # 遍历view_points_dic 取出 当前时刻的数据
        for key, view_points in view_points_dic.items():
            # print(key, view_points)
            #  获取监测点在模型中的坐标
            coordinates_list = view_points.coordinates_list
            english_name_list = view_points.english_column_name_list
            if coordinates_list is not None and english_name_list is not None:
                try:
                    data_list = np.array(data_df[view_points.column_name_list])[-1]
                except BaseException as e:
                    data_list = np.array([0] * len(view_points.column_name_list))
                # 格式化为两位小数
                formatted_data_list = [f'{data_e:.2f}' for data_e in data_list]
                local_labels_list = np.array(
                    [english_name + ':' + str(data) for english_name, data in zip(english_name_list, formatted_data_list)])
                now_labels_list = np.concatenate((now_labels_list, local_labels_list))
        now_labels_list = list(now_labels_list)
        # print('now_labels_list', now_labels_list)
        # now_labels_list = [str((i + 1) * np.random.randint(1, 20)) for i in range(len(labels_coordinates_list))]
        # TODO 更新当前云图的监测点显示数据
        # window.update_plotter_now_labels(labels_coordinates_list, now_labels_list)
        # 此处不再直接操作主界面，而是发射信号给MySignals函数中的text_print，
        # 传递参数包括要操作哪个位置和要操作的内容
        window.ms.text_print.emit(labels_coordinates_list, now_labels_list)

        global pickPoint, pickPoint_index
        # 如果 有选中的点，则传递参数包括要操作哪个位置和要操作的内容
        if pickPoint is not None and pickPoint_index != 0:
            # 传递参数包括要操作哪个位置和要操作的内容
            window.ms.pickPointLabel.emit(pickPoint, pickPoint_index)
        # lock_global.release()   # 开锁
        # print("read...开锁")
        time.sleep(1)

# 写入线程  每秒写入一行实验数据，模拟真实实验一秒一行
def write_row_by_second():
    # 初始化计数器
    row_to_copy = 1
    print(f"total_rows={total_rows}")
    while row_to_copy <= total_rows:
        lock_file.acquire()  # 锁上
        print("我正在写入，上锁")
        # 打开目标文件
        target_workbook = openpyxl.load_workbook(target_file_path)
        target_sheet = target_workbook.active  # 假设目标文件只有一个工作表

        # 复制数据，每次复制一行,,所有列
        for col in range(1, source_sheet.max_column + 1):
            cell_value = source_sheet.cell(row=row_to_copy, column=col).value
            target_sheet.cell(row=row_to_copy, column=col, value=cell_value)

        # 保存目标文件
        target_workbook.save(target_file_path)
        target_workbook.close()

        # 打印信息
        print(f"Copied data from row {row_to_copy}: {source_sheet[row_to_copy]}")
        lock_file.release()  # 开锁
        print("我写入完毕，开锁")
        # 增加行计数
        row_to_copy += 1

        if row_to_copy <= total_rows:
            # 1秒复制一行
            time.sleep(1)

    # 关闭源文件的工作簿
    source_workbook.close()




# 写入vtu文件线程  降温3分钟后，开始每秒将当前实验时刻的信息（拟合后的） 存入vtu文件
def write_vtu_by_second():
    is_first = True  # 判断是否为第一次循环，用于添加 右键 拾取点
    while True:
        global end_time, time_list, strain_lists, vtuModel, vtu_storage_directory, test_Temperature_lists,\
            now_vtu_filename, view_points_dic, real_time_list
        event_break.wait()  # TODO 断裂后，阻塞线程
        print("write_vtu thread waiting...")
        event.wait()  # TODO 开始降温  3分钟后 阻塞放开
        print("write_vtu thread going...")

        # time.sleep(1)
        # continue

        # TODO 每秒修改vtu文件名为当前时刻
        time_path = real_time_list[-1].replace("/", "-").replace(":", "：")
        now_vtu_filename = f"{vtu_storage_directory}/CylinderModelTime_{time_path}.vtu"
        vtuModel.set_vtu_filename(now_vtu_filename)

        # TODO 把数据拟合并插值所有数据，设置到vtuModel的节点数据中 <温度>   BDFAndTemperatureToVtuText6.py中的方法
        T_know_points = view_points_dic.get("温度").coordinates_list
        T_now_temperatures_list = view_points_dic.get("温度").data[-1, :]
        eFunFitting_to_vtuModel(vtuModel, T_know_points, 'Temperature', T_now_temperatures_list, number=1)

        # TODO 用给定的指数函数拟合圆柱时可能会出现：拟合不准确导致报错
        is_error = False
        try:
            # 把数据拟合并插值所有的应变、应力并存入vtuModel对象中
            strain_stress_to_vtu_1HZ(vtuModel, view_points_dic)  # TODO BDFAndTemperatureToVtuText6.py中的方法
            # vtuModel.write_to_vtu()
        except BaseException as e:
            print(e)
            is_error = True
            now_vtu_filename = f"{vtu_storage_directory}/CylinderModelTime_{time_path}_CloudFittingError.vtu"
            vtuModel.set_vtu_filename(now_vtu_filename)
            # QMessageBox.warning(None, '警告', '圆柱模型公式拟合出错')
        else:
            # 没有出现错误
            pass
        finally:
            vtuModel.write_to_vtu()
            print(f"生成的vtu文件：{now_vtu_filename}")
            print(f"是否发生错误：{is_error}")
            if not is_error:  # 没有出现错误
                # 刷新当前时刻的云图
                print("执行刷新云图.....")
                window.update_plotter_now(now_vtu_filename)

                # TODO 用于Pyvista 绘制曲线图
                T_column_index = 0
                r_strain_column_index = 0
                theta_strain_column_index = 0
                z_strain_column_index = 0

                # 点击 事件
                def cloud_now_pick_callback(point):
                    global pickPoint, pickPoint_index
                    points = window.mesh_now.points
                    # 计算点与数组中每个点的距离
                    distances = np.linalg.norm(points - point, axis=1)
                    # 找到距离最接近的点的索引
                    index = np.argmin(distances)
                    print(f"index={index}")
                    point = points[index]
                    pickPoint = point
                    pickPoint_index = index
                    node_data_name = window.ui.cbox_predict_dataType.currentText()  # 获取选择的数据类型
                    node_data = window.mesh_now.point_data[node_data_name]
                    # print(f"cloud_now_pick_callback[node_data_name]{node_data_name}")
                    # print(f"cloud_now_pick_callback[node_data]{node_data}")
                    # mesh = pv.Cube(center=point, x_length=1, y_length=1, z_length=1)
                    # window.plotter_now[0,0].add_mesh(mesh, name='cloud_now_picked', style='wireframe', color='r')
                    window.plotter_now[0, 0].add_point_labels(pickPoint, [f"{node_data[pickPoint_index]:.2f}"],
                                                              name='cloud_now_picked',
                                                              italic=True,
                                                              text_color='black',
                                                              # point_color='red',
                                                              show_points=False,
                                                              font_size=22,
                                                              bold=True,
                                                              font_family='times',  # courier times  arial
                                                              shape='rounded_rect',  # rect（矩形）  rounded_rect（圆角矩形）
                                                              shape_color='yellow',  # grey
                                                              shape_opacity=0.3,  # 透明度 0-1
                                                              margin=10,  # 边距
                                                              always_visible=True,
                                                              )
                    nonlocal T_column_index, r_strain_column_index, theta_strain_column_index, z_strain_column_index
                    # 获取 监测点对象
                    T_view_point = view_points_dic.get("温度")
                    r_strain_view_point = view_points_dic.get("r方向应变")
                    theta_strain_view_point = view_points_dic.get("top_theta方向应变")
                    z_strain_view_point = view_points_dic.get("z方向应变")
                    # 获取监测点数据 列 数量  用于随机
                    T_shape = T_view_point.data.shape[1]
                    r_strain_shape = r_strain_view_point.data.shape[1]
                    theta_strain_shape = theta_strain_view_point.data.shape[1]
                    z_strain_shape = z_strain_view_point.data.shape[1]
                    # 随机选择一个监测点
                    T_column_index = np.random.randint(0, T_shape)
                    r_strain_column_index = np.random.randint(0, r_strain_shape)
                    theta_strain_column_index = np.random.randint(0, theta_strain_shape)
                    z_strain_column_index = np.random.randint(0, z_strain_shape)
                    window.strain_chart.visible = True  # 显示绘图
                    window.temperature_chart.visible = True  # 显示绘图

                window.plotter_now[0, 0].enable_surface_point_picking(callback=cloud_now_pick_callback,
                                                                      show_point=False,
                                                                      show_message=False) if is_first else None
                is_first = False  # 第一次执行完，置为False
                # 更新Pyvista曲线图
                window.update_pyvista_charts(time_list[:], T_column_index, r_strain_column_index,
                                             theta_strain_column_index, z_strain_column_index)

            time.sleep(1)





# 预测绘图 线程
def plot_run_by_second():
    # is_first = True    # 判断是否为第一次循环，用于添加 右键 拾取点
    while True:
        # end_time 是当前实验时刻，[降温开始，end_time]用于预测的训练数据
        global end_time, time_list, strain_lists, test_Temperature_lists, decrease_index, real_time_list, ax3_strain_legend_2
        event_break.wait()  # TODO 断裂后，阻塞线程

        print("plot thread waiting...")
        event.wait()    # TODO 开始降温  3分钟后 阻塞放开
        print("plot thread going...")

        ''''''
        # print(f"time_list[:]======={time_list[:]}")
        # print(f"view_points_dic.get('温度').data[:, 0]======={view_points_dic.get('温度').data[:, 0]}")
        # window.temperature_line1.update(time_list[:], view_points_dic.get("温度").data[:, 0])
        # print(f"real_time_list[decrease_index:]>>>>>>>{real_time_list[decrease_index:]}")

        # 遍历view_points_dic字典
        for view_type in view_points_dic.keys():
            e_view_points = view_points_dic.get(view_type)
            if e_view_points is not None and "温度" in e_view_points.monitoring_type:
                # TODO 获取预测的未来数据，并存入到监测点的对象中
                get_predict_data_to_view_points(e_view_points, type=view_type)
                # TODO 绘制曲线图
                window.update_temperature_strain_charts(real_time_list[decrease_index:], e_view_points, "温度")
            if e_view_points is not None and "应变" in e_view_points.monitoring_type:
                # TODO 获取预测的未来数据，并存入到监测点的对象中
                get_predict_data_to_view_points(e_view_points, type=view_type)
                # TODO 绘制曲线图
                window.update_temperature_strain_charts(real_time_list[decrease_index:], e_view_points, "应变")

        window.ax_temperature1.legend(loc='lower left',  # 设置图例位置
                                      labelspacing=0,  # 设置图例间距
                                      handlelength=2,  # 设置图例中线的长度
                                      ncol=2,  # 设置图例的列数
                                      fontsize=8,  # 设置图例字体大小
                                      # labelcolor='white',
                                      shadow=True,  # 设置图例阴影
                                      # edgecolor="None",     # 设置图例边框颜色的参数, None 没有边框
                                      # frameon=False,    # 是否显示边框
                                      draggable=True  # 设置图例可拖动
                                      ) if window.ax_temperature1.get_legend() is None else None
        window.ax_temperature2.legend(loc='lower left',  # 设置图例位置
                                      labelspacing=0,  # 设置图例间距
                                      handlelength=2,  # 设置图例中线的长度
                                      ncol=2,  # 设置图例的列数
                                      fontsize=8,  # 设置图例字体大小
                                      # labelcolor='white',
                                      shadow=True,  # 设置图例阴影
                                      # edgecolor="None",  # 设置图例边框颜色的参数, None 没有边框
                                      # frameon=False,  # 是否显示边框
                                      draggable=True  # 设置图例可拖动
                                      ) if window.ax_temperature2.get_legend() is None else None
        strain1_lines_list = np.array(list(window.ax_strain1_lines.values()))
        filtered_handles = [art for art in strain1_lines_list if not art.get_label().startswith('_')]
        if ax3_strain_legend_2 is None:
            ax3_strain_legend_2 = window.ax_strain1.legend(loc='upper left',  # 设置图例位置
                                                      handles=filtered_handles,
                                                      labelspacing=0,  # 设置图例间距
                                                      handlelength=2,  # 设置图例中线的长度
                                                      ncol=2,  # 设置图例的列数
                                                      fontsize=8,  # 设置图例字体大小
                                                      # labelcolor='white',
                                                      shadow=True,  # 设置图例阴影
                                                      # edgecolor="None",  # 设置图例边框颜色的参数, None 没有边框
                                                      # frameon=False,  # 是否显示边框
                                                      draggable=True  # 设置图例可拖动
                                                      )
            window.ax_strain1.add_artist(ax3_strain_legend_2)
        window.ax_strain2.legend(loc='upper left',  # 设置图例位置
                                      labelspacing=0,  # 设置图例间距
                                      handlelength=2,  # 设置图例中线的长度
                                      ncol=2,  # 设置图例的列数
                                      fontsize=8,  # 设置图例字体大小
                                      # labelcolor='white',
                                      shadow=True,  # 设置图例阴影
                                      # edgecolor="None",  # 设置图例边框颜色的参数, None 没有边框
                                      # frameon=False,  # 是否显示边框
                                      draggable=True  # 设置图例可拖动
                                      ) if window.ax_strain2.get_legend() is None else None
        break_time_lists = []
        break_strain_lists = []
        # 遍历view_points_dic字典
        for view_type in view_points_dic.keys():
            e_view_points = view_points_dic.get(view_type)
            if e_view_points is not None and "应变" in e_view_points.monitoring_type:
                # if e_view_points.predict_data
                break_time_lists.extend(e_view_points.break_time_list)
                break_strain_lists.extend(e_view_points.break_data_list)

        break_time_lists = np.array(break_time_lists)
        break_strain_lists = np.array(break_strain_lists)
        # 判断最早断裂时间和应变
        if len(break_time_lists) > 0:
            earlier_break_index = np.where(break_time_lists == np.min(break_time_lists))[0]
            earlier_break_index = earlier_break_index[0]
            # print(f"earlier_break_index = {earlier_break_index}")
            earlier_break_index = int(earlier_break_index)
            earlier_break_time = break_time_lists[earlier_break_index]
            earlier_break_strain = break_strain_lists[earlier_break_index]
            # print(f"earlier_break_time = {earlier_break_time}")
            # print(f"earlier_break_strain = {earlier_break_strain}")
            set_text(earlier_break_time, earlier_break_strain)
        else:
            plt_title.set_text('')



        # 异步刷新绘图，不会阻塞
        # fig.canvas.draw_idle()
        # 立即刷新绘图， 如果鼠标不在Axis中
        # fig.canvas.draw() if not mouse_in_axis else None
        try:
            fig.canvas.draw() if not mouse_in_axis else None
            # 使用 mplcursors 注册悬停事件处理函数
            # sel.annotation.arrow_patch.set_edgecolor('white')设置箭头颜色
            # mplcursors.cursor().connect("add", lambda sel: sel.annotation.arrow_patch.set_edgecolor('white'))
        except BaseException as e:
            print(f"1hz flash fig.canvas.draw() error：{e}")
        '''
        # TODO 用于Pyvista 绘制曲线图
        T_column_index = 0
        r_strain_column_index = 0
        theta_strain_column_index = 0
        z_strain_column_index = 0
        # 点击 事件
        def cloud_now_pick_callback(point):
            global pickPoint, pickPoint_index
            points = window.mesh_now.points
            # 计算点与数组中每个点的距离
            distances = np.linalg.norm(points - point, axis=1)
            # 找到距离最接近的点的索引
            index = np.argmin(distances)
            print(f"index={index}")
            point = points[index]
            pickPoint = point
            pickPoint_index = index
            node_data_name = window.ui.cbox_predict_dataType.currentText()  # 获取选择的数据类型
            node_data = window.mesh_now.point_data[node_data_name]
            print(f"cloud_now_pick_callback[node_data_name]{node_data_name}")
            print(f"cloud_now_pick_callback[node_data]{node_data_name}")
            # mesh = pv.Cube(center=point, x_length=1, y_length=1, z_length=1)
            # window.plotter_now[0,0].add_mesh(mesh, name='cloud_now_picked', style='wireframe', color='r')
            window.plotter_now[0, 0].add_point_labels(pickPoint, [f"{node_data[pickPoint_index]:.2f}"],
                                                    name='cloud_now_picked',
                                                    italic=True,
                                                    text_color='black',
                                                    # point_color='red',
                                                    show_points=False,
                                                    font_size=22,
                                                    bold=True,
                                                    font_family='times',  # courier times  arial
                                                    shape='rounded_rect',  # rect（矩形）  rounded_rect（圆角矩形）
                                                    shape_color='yellow',  # grey
                                                    shape_opacity=0.3,  # 透明度 0-1
                                                    margin=10,  # 边距
                                                    always_visible=True,
                                                    )
            nonlocal T_column_index, r_strain_column_index, theta_strain_column_index, z_strain_column_index
            # 获取 监测点对象
            T_view_point = view_points_dic.get("温度")
            r_strain_view_point = view_points_dic.get("r方向应变")
            theta_strain_view_point = view_points_dic.get("top_theta方向应变")
            z_strain_view_point = view_points_dic.get("z方向应变")
            # 获取监测点数据 列 数量  用于随机
            T_shape = T_view_point.data.shape[1]
            r_strain_shape = r_strain_view_point.data.shape[1]
            theta_strain_shape = theta_strain_view_point.data.shape[1]
            z_strain_shape = z_strain_view_point.data.shape[1]
            # 随机选择一个监测点
            T_column_index = np.random.randint(0, T_shape)
            r_strain_column_index = np.random.randint(0, r_strain_shape)
            theta_strain_column_index = np.random.randint(0, theta_strain_shape)
            z_strain_column_index = np.random.randint(0, z_strain_shape)
            window.strain_chart.visible = True  # 显示绘图
            window.temperature_chart.visible = True  # 显示绘图

        window.plotter_now[0,0].enable_surface_point_picking(callback=cloud_now_pick_callback, show_point=False,
                                                             show_message=False) if is_first else None
        is_first = False  # 第一次执行完，置为False
        # 更新Pyvista曲线图
        window.update_pyvista_charts(time_list[:], T_column_index, r_strain_column_index, theta_strain_column_index, z_strain_column_index)
        '''
        time.sleep(1)
        continue


        # lock_global.acquire()  # 上锁
        # print("plot...上锁")
        # if event.is_set():  # 判断是否设置了flag
        #     print("plot sleep...")
        #     time.sleep(1)   # 没有设置，一直睡眠
        # else:

        # lock_global.release()  # 开锁
        # print("plot...开锁")

# 获取预测的未来数据，并存入到监测点的对象中
def get_predict_data_to_view_points(e_view_points, type):
    shape = e_view_points.data.shape[1]
    pre_data_lists = []  # 预测数据矩阵
    break_time = []
    break_strain = []
    for i in range(shape):
        e_data_column = e_view_points.data[:, i]
        time_list_local = time_list[:]
        # print(f"len(e_data_column)>>>>>>>{len(e_data_column)}")
        # print(f"len(time_list_local)>>>>>>>{len(time_list_local)}")
        # 防止时间列表和数据列表长度不一致
        if len(time_list_local) > len(e_data_column):
            time_list_local = time_list_local[:len(e_data_column)]
        elif len(time_list_local) < len(e_data_column):
            e_data_column = e_data_column[:len(time_list_local)]
        predict_time_list, predict_temperature_list, break_time_list, break_strain_list = predict_data_func(time_list_local, e_data_column, type)
        e_view_points.predict_time = predict_time_list
        pre_data_lists.append(predict_temperature_list)
        break_time.extend(break_time_list)
        break_strain.extend(break_strain_list)
    pre_data_lists = np.array(pre_data_lists)
    pre_data_lists = pre_data_lists.T  # 转置一下
    e_view_points.predict_data = pre_data_lists  # 添加到监测点字典中
    break_time = np.array(break_time)
    break_strain = np.array(break_strain)
    e_view_points.break_time_list = break_time
    e_view_points.break_data_list = break_strain




# 定义多项式函数，用于预测
def exponential_func_T(x, a, b):
    # return a * x**2 + b * x + c
    return a * x + b

# 未来温度、应变预测
def predict_data_func(time_list, data_list, type):
    exponential_function = None
    if "温度" in type:
        exponential_function = exponential_func_T
    elif "应变" in type:
        exponential_function = exponential_func_strain
    else:
        print("type error")
    time_list_local = time_list[:]
    data_list_local = data_list[:]
    end_time_local = time_list_local[-1]
    # 拆分数据集
    X_train, X_test, y_train, y_test = train_test_split(time_list_local, data_list_local, test_size=0.2, random_state=0)

    # 拟合指数模型
    params, covariance = curve_fit(exponential_function, X_train, y_train)

    # 进行预测
    y_pred = exponential_function(X_test, *params)
    mse = mean_squared_error(y_test, y_pred)
    # print('均方误差:', mse)
    # TODO 预测多久 单位秒
    future_time = np.array([time for time in range(end_time_local, end_time_local + 60 * 1)])
    # print("预测时间:", future_time)
    future_data = exponential_function(future_time, *params)
    # print('未来数据预测:', future_data)
    break_time_list = []
    break_strain_list = []
    if "应变" in type:
        for  i in range(len(future_data)):
            if future_data[i] > 520 and future_data[i] < 660:
                break_time_list.append(future_time[i])
                break_strain_list.append(future_data[i])



    return future_time, future_data, break_time_list, break_strain_list


# 定义多项式函数，用于预测
def exponential_func_strain(x, a, b, c, d):
    # return a * x**2 + b * x + c
    return a * x**3 + b * x**2 + c * x + d

# 判断什么时间开始降温
def get_start_time(time_list):
    # global bg_Temperature_list
    local_bg_Temperature_list = None
    try:
        # 判断view_points_dic.get("环境温度").data列数是否>1
        if view_points_dic.get("环境温度").data.shape[1] > 1:
            print(f"==========================={view_points_dic.get('环境温度').data.shape[1]}")
        local_bg_Temperature_list = view_points_dic.get("环境温度").data[:,0]
        # print(f"local_bg_Temperature_list= {local_bg_Temperature_list}")
    except:
        print("获取环境温度失败")
    decrease_index = 0
    start_time_of_decrease = 0
    for i in range(1, len(local_bg_Temperature_list) - 3):
        # if local_bg_Temperature_list[i] - local_bg_Temperature_list[i + 3] >= 1:
        if local_bg_Temperature_list[i] - local_bg_Temperature_list[i + 3] >= 0.1:
            # print(local_bg_Temperature_list[i])
            # print(local_bg_Temperature_list[i + 3])
            decrease_index = i + 2
            start_time_of_decrease = time_list[decrease_index]
            break  # 找到第一个下降后就退出循环
    # TODO 未找到下降点，返回 0
    return start_time_of_decrease, decrease_index

# 应变预测  return [time_list, strain_list, future_time, future_strain, break_time_list, break_strain_list]
def predict_1HZ_flush(time_list, strain_list, start_time_of_decrease, end_time):
    # print(f"end_time=====>>>>>{end_time}")
    # start_index = np.where(time_list == start_time_of_decrease)[0][0]
    start_index = np.where(time_list == 0)[0][0]    # TODO 因为归一化处理，所以开始降温时刻，为0
    end_index = np.where(time_list == end_time)[0][0]
    # print(f"start_index======>>>{start_index}\nend_index=======>>>>{end_index}")
    time_list_local = time_list[start_index: end_index]
    strain_list_local = strain_list[start_index: end_index]

    # 拆分数据集
    X_train, X_test, y_train, y_test = train_test_split(time_list_local, strain_list_local, test_size=0.2, random_state=0)

    # 拟合指数模型
    params, covariance = curve_fit(exponential_func_strain, X_train, y_train)
    # params, covariance = curve_fit(exponential_func, time, strain1)

    # a_fit, b_fit, c_fit, d_fit = params
    # 输出拟合参数
    # print("拟合参数 a:", a_fit)
    # print("拟合参数 b:", b_fit)
    # print("拟合参数 c:", c_fit)
    # print("拟合参数 d:", d_fit)

    # 进行预测
    y_pred = exponential_func_strain(X_test, *params)
    mse = mean_squared_error(y_test, y_pred)
    print('均方误差:', mse)

    future_time = np.array([time for time in range(end_time, end_time + 60 * 1)])
    print("预测时间:", future_time)
    future_strain = exponential_func_strain(future_time, *params)
    print('未来应变预测:', future_strain)
    # 判断应变何时 达到断裂判断的范围
    break_time_list = []
    break_strain_list = []
    for i in range(len(future_strain)):
        if future_strain[i] > 520 and future_strain[i] < 660:
            # print(future_strain[i])
            break_time_list.append(future_time[i])
            break_strain_list.append(future_strain[i])
    break_time_list = np.array(break_time_list)
    break_strain_list = np.array(break_strain_list)
    print('应变达到断裂的时间范围:')
    print(break_time_list)


    return [time_list, strain_list, future_time, future_strain, break_time_list, break_strain_list]


def get_data_from_excel(file_name, columns_name_list=None):
    # 使用上下文管理器读取 Excel 文件
    with pd.ExcelFile(file_name) as file:
        data_df = pd.read_excel(file, sheet_name="Sheet1")  # 请替换成你的工作表名
    # data_df = pd.read_excel(file_name)
    my_data_df = data_df.loc[:, columns_name_list]
    return np.array(my_data_df)

# 预测曲线的提示文字
def set_text(break_time, break_strain):
    # text_pt.set_position((break_time_list[0], break_strain_list[0]-100))
    text_pt.set_text("应变值：%.3f" % break_strain)
    # text_pt.set_text("最早时间：%d\n应变值：%.3f" % (break_time, break_strain))
    text_worn.set_text('警告：%d秒后可能发生断裂' % (break_time - end_time))
    plt_title.set_text(f'警告：{(break_time - end_time)}秒后可能发生断裂 应变值：{break_strain:.3f}')
    print("警告：  {}秒后出现应变超过+520，可能出现断裂".format(break_time - end_time))




# 鼠标拖动 处理事件
def call_move(event):
    # print(event.name)
    global mPress, startx, starty, mouse_in_axis
    mouse_x = event.x
    mouse_y = event.y
    axtemp = event.inaxes
    if event.name == 'button_press_event':
        if axtemp and event.button == 1:
            # 判断拖动的是否为图例
            if axtemp.get_legend():
                legend_bbox = axtemp.get_legend().get_window_extent()
                left_bottom = legend_bbox.get_points()[0]
                right_top = legend_bbox.get_points()[1]

                if left_bottom[0] <= mouse_x <= right_top[0] and left_bottom[1] <= mouse_y <= right_top[1]:
                    # print("在图例上按下鼠标")
                    # 在图例上按下鼠标
                    mPress = False
                    return
            # 没有图例的情况
            # print("在 Axes 上按下鼠标")
            # 在 Axes 上按下鼠标
            mPress = True
            startx = event.xdata
            starty = event.ydata
            return
    elif event.name == 'button_release_event':
        if axtemp and event.button == 1:
            mPress = False
    elif event.name == 'motion_notify_event':
        # 判断鼠标是否在Axis上
        mouse_in_axis = True if axtemp is not None else False
        if axtemp and event.button == 1 and mPress:
            if axtemp.get_legend():
                legend_bbox = axtemp.get_legend().get_window_extent()
                left_bottom = legend_bbox.get_points()[0]
                right_top = legend_bbox.get_points()[1]

                if left_bottom[0] <= mouse_x <= right_top[0] and left_bottom[1] <= mouse_y <= right_top[1]:
                    print("在图例上移动鼠标")
                    # 在图例上按下鼠标
                    mPress = False
                    return

            # 没有图例的情况
            # print("在Axes上移动鼠标")
            x_min, x_max = axtemp.get_xlim()
            y_min, y_max = axtemp.get_ylim()
            w = x_max - x_min
            h = y_max - y_min
            # print(event)
            # 移动
            mx = event.xdata - startx
            my = event.ydata - starty
            # 注意这里， -mx,  因为下一次 motion事件的坐标，已经是在本次做了移动之后的坐标系了，所以要体现出来
            # startx=event.xdata-mx  startx=event.xdata-(event.xdata-startx)=startx, 没必要再赋值了
            # starty=event.ydata-my
            # print(mx,my,x_min,y_min,w,h)
            axtemp.set(xlim=(x_min - mx, x_min - mx + w))
            axtemp.set(ylim=(y_min - my, y_min - my + h))
            # fig.canvas.draw_idle()  # 绘图动作实时反映在图像上
            try:
                fig.canvas.draw_idle()
            except BaseException as e:
                print(f"鼠标拖动axis 处理事件 fig.canvas.draw_idle() error：{e}")

    return

# 滚轮滚动 处理事件
def call_scroll(event):
    # print(event.name)
    axtemp = event.inaxes
    # print('event:', event)
    # print(event.xdata, event.ydata)
    # 计算放大缩小后， xlim 和ylim
    if axtemp:
        x_min, x_max = axtemp.get_xlim()
        y_min, y_max = axtemp.get_ylim()
        w = x_max - x_min
        h = y_max - y_min
        curx = event.xdata
        cury = event.ydata
        curXposition = (curx - x_min) / w
        curYposition = (cury - y_min) / h
        if event.button == 'down':
            # print('befor:', w, h)
            w = w * 1.1
            h = h * 1.1
            # print('down', w, h)
        elif event.button == 'up':
            # print('befor:', w, h)
            w = w / 1.1
            h = h / 1.1
            # print('up', w, h)
        # print(curXposition, curYposition)
        newx = curx - w * curXposition
        newy = cury - h * curYposition
        axtemp.set(xlim=(newx, newx + w))
        # axtemp.set_xticklabels(text_time_list)
        axtemp.set(ylim=(newy, newy + h))
        fig.canvas.draw_idle()  # 绘图动作实时反映在图像上
        try:
            fig.canvas.draw_idle()
        except BaseException as e:
            print(f"滚轮滚动 处理事件 fig.canvas.draw_idle() error：{e}")



if __name__ == "__main__":
    # global data_excel, bg_Temperature_list, time_list, strain_lists, start_time_of_decrease, end_time, columns, colors,\
    #             lines_list
    data_excel = []
    app = QApplication(sys.argv)

    plt.rcParams["font.family"] = "Microsoft YaHei"

    # 设置图表的背景颜色为灰色
    # plt.gcf().set_facecolor('red')

    # 创建初始的图形
    fig, axs = plt.subplots(nrows=2, ncols=2, figsize=(16, 8))
    columns = ['应变1/ue', '应变2/ue', '应变3/ue', '应变4/ue',
               '应变5/ue', '应变6/ue', '应变7/ue', '应变8/ue',
               '应变9/ue', '应变10/ue', '应变11/ue', '应变12/ue',
               '应变13/ue', '应变14/ue', '应变15/ue', '应变16/ue']
    colors = ['blue', 'green', 'orange', 'purple',
              'brown', 'pink', 'gray', '#00A0A0',
              '#000080', '#FFA500', '#FF00FF', '#008080',
              '#A0A080', '#FD05A0', '#CD008F', '#B08D80']
    ax = axs[0,0]
    ax2 = axs[0,1]
    ax3 = axs[1,0]
    legent_default1, = ax3.plot([], [], label="历史变化", color='black', ls='-')
    legent_default2, = ax3.plot([], [], label="未来趋势", color='black', ls='--')
    ax_strain_legend_1 = ax3.legend(handles=[legent_default1, legent_default2],
               bbox_to_anchor=(1.03, 1.27),
               loc='upper left',
               labelspacing=0,  # 设置图例间距
               handlelength=2,  # 设置图例中线的长度
               shadow=True,  # 设置图例阴影
               draggable=True,  # 设置图例可拖动
               borderaxespad=0.
               )
    ax3.add_artist(ax_strain_legend_1)
    # 统一设置颜色属性
    # 获取当前图表对象(fig)，并设置背景颜色为黑色
    plt.gcf().set_facecolor('#23262f')  # 31363b
    now_time = datetime.now()
    for i in range(len(axs)):
        for j in range(len(axs[i])):
            ax_i = axs[i, j]
            ax_i.grid(ls=':', color='white')  # 设置网格，颜色为白色
            # 设置坐标轴的背景颜色为灰色gray
            ax_i.set_facecolor('#23262f')  # 232629
            # 设置坐标轴刻度、数值和标题的颜色为白色
            ax_i.tick_params(axis='both', colors='white')
            ax_i.title.set_color('white')
            # 获取坐标轴的边框对象并设置颜色为白色
            for spine in ax_i.spines.values():
                spine.set_edgecolor('white')


            # # 启用坐标轴的次要刻度线
            # ax.minorticks_on()
            # # 配置坐标轴的刻度
            ax_i.tick_params(axis='both', which='both', direction='in', bottom=True, top=True, left=True, right=True)
            ax_i.xaxis.set_major_formatter(mdate.DateFormatter('%H:%M:%S'))  # 设置时间标签显示格式
            ax_i.set_xlim(now_time, now_time + timedelta(seconds=4))
            ax_i.margins(0)  # 调整坐标轴两端的空白
            ax_i.set_xlabel('时间', color='white', fontsize=14)
            ax_i.set_ylabel('应变(um/m)', color='white', fontsize=14)
            plt.sca(ax_i)
            # ax_i.set_xticklabels(rotation=45)
            # ax_i.xticks(rotation=30)
            plt.xticks(rotation=30)
    # for ax_i in axs.flatten():
    #     plt.xticks(rotation=30)
    ax.set_ylabel('温度', color='white')
    ax2.set_ylabel('温度', color='white')
    text_time_list = ['15:33:22', '15:33:23', '15:33:24', '15:33:25', '15:33:26',]
    # 使用AutoLocator自动选择刻度位置
    # ax.xaxis.set_major_locator(AutoLocator())
    # ax.yaxis.set_major_locator(AutoLocator())
    # 使用MultipleLocator设置x轴刻度间隔为100
    # ax.xaxis.set_major_locator(MultipleLocator(1))
    # ax.yaxis.set_major_locator(MultipleLocator(100))
    # 使用MaxNLocator设置x轴刻度最多显示5个
    # ax.xaxis.set_major_locator(MaxNLocator(5))
    # fig.autofmt_xdate()
    # plt.xticks(rotation=30)


    # ax.set_xlabel('时间(s)', color='white')
    # ax.set_ylabel('应变(um/m)', color='white')
    # fig.autofmt_xdate()

    # 设置坐标轴范围
    # ax.set_xlim(0, 1000)
    # ax.set_ylim(-200, 800)
    # ax.plot([], [], label="训练数据", color='black', ls='-')
    # ax.plot([], [], label="预测数据", color='black', ls='--')
    # ax.plot([], [], label="可能断裂点", color='red', ls=':')
    lines_list = []
    # for i in range(len(columns)):
    #     line, = ax.plot([], [], label=f'应变{i + 1}', color=colors[i], ls='-')
    #     line_future, = ax.plot([], [], color=colors[i], ls='--')
    #     line_break, = ax.plot([], [], color='red', ls=':')
    #     # print(type(line))
    #     lines_list.extend([line, line_future, line_break])

    # 添加图例
    # ax.legend(loc=4, labelspacing=1, handlelength=2, fontsize=8, shadow=True)
    # ax.legend(loc='lower right', labelspacing=0, handlelength=2, ncol=4, fontsize=8, shadow=True)
    text_pt = ax.text(ax.get_xlim()[0], ax.get_ylim()[1], '', color='white', fontsize=14)
    text_worn = ax.text(ax.get_xlim()[0], ax.get_ylim()[1], '', color='red', fontsize=14)
    plt_title = plt.suptitle("", x=0.5, y=0.95, color='red', fontsize=14)
    # title.set_text("hhhhhh")

    # 使用 tight_layout 自动调整 Figure 和 Axes 之间的间距
    # plt.tight_layout()
    # 调整 Figure 和 Axes 之间的间距
    plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.13)
    fig.subplots_adjust(hspace=0.4)
    fig.subplots_adjust(wspace=0.4)

    # 添加鼠标事件
    fig.canvas.mpl_connect('scroll_event', call_scroll)
    fig.canvas.mpl_connect('button_press_event', call_move)
    fig.canvas.mpl_connect('button_release_event', call_move)
    # fig.canvas.mpl_connect('draw_event', call_move)
    fig.canvas.mpl_connect('motion_notify_event', call_move)


    # TODO 创建窗体对象
    window = MyWindow(fig)

    # 设置窗体最大化显示
    window.ui.showMaximized()

    # 打开源文件和目标文件
    # source_file_path = "../CylinderData/综合20231023150155_3-columns_reset.xlsx"
    source_file_path = "../CylinderData/仿真数据（改）.xlsx"
    # target_file_path = None

    # 获取源文件的总行数
    source_workbook = openpyxl.load_workbook(source_file_path)
    source_sheet = source_workbook.active  # 假设源文件只有一个工作表
    total_rows = source_sheet.max_row

    lock_file = threading.Lock()
    lock_global = threading.Lock()
    event = threading.Event()   # 用于 降温 x分钟后 放行 预测线程和写入vtu线程
    event_break = threading.Event()     # 用于 断裂后 阻塞 读取线程、预测线程、写入vtu线程
    event_break.set()  # 初始 设为绿灯（True），放行 读取线程、预测线程、写入vtu线程
    t_read = threading.Thread(target=read_row_by_second, )
    t_read.setDaemon(True)

    t_write = threading.Thread(target=write_row_by_second, )
    t_write.setDaemon(True)

    t_write_vtu = threading.Thread(target=write_vtu_by_second, )
    t_write_vtu.setDaemon(True)

    t_plot = threading.Thread(target=plot_run_by_second, )
    t_plot.setDaemon(True)

    t_text = threading.Thread(target=thread_text, )
    t_text.setDaemon(True)



    # setup stylesheet
    apply_stylesheet(app, theme='dark_blue.xml')
    '''
    ['dark_amber.xml',
 'dark_blue.xml',
 'dark_cyan.xml',
 'dark_lightgreen.xml',
 'dark_pink.xml',
 'dark_purple.xml',
 'dark_red.xml',
 'dark_teal.xml',
 'dark_yellow.xml',
 'light_amber.xml',
 'light_blue.xml',
 'light_cyan.xml',
 'light_cyan_500.xml',
 'light_lightgreen.xml',
 'light_pink.xml',
 'light_purple.xml',
 'light_red.xml',
 'light_teal.xml',
 'light_yellow.xml']
    '''




    window.ui.show()

    sys.exit(app.exec_())