import sys
import threading
import time

import openpyxl
import pandas as pd
from PySide2.QtWidgets import QApplication, QWidget, QFileDialog
from PySide2.QtUiTools import QUiLoader
from PySide2.QtCore import QFile, QIODevice

sourceFilePath = ""     # 源文件路径
targetFilePath = ""     # 目标文件路径
source_data = None      # 源文件数据


class MyWindow:
    def __init__(self):
        # 从文件中加载UI定义
        self.load_ui()
        # 绑定事件
        # 给 选择源文件 按钮 绑定事件
        self.ui.btn_selectSourceFile.clicked.connect(self.click_selectSourceFile)
        # 给 选择目标文件 按钮 绑定事件
        self.ui.btn_selectTargetFile.clicked.connect(self.click_selectTargetFile)
        # 给 开始复制 按钮 绑定事件
        self.ui.btn_runCopy.clicked.connect(self.click_runCopy)
        # 连接文本编辑器的文本变化信号到滚动到底部的槽
        self.ui.textEdit_log.textChanged.connect(self.scroll_to_bottom)

    def load_ui(self):
        # 从文件中加载UI定义
        ui_file_name = './UI/WriteMain.ui'
        ui_file = QFile(ui_file_name)
        if not ui_file.open(QIODevice.ReadOnly):
            print(f"Cannot open {ui_file_name}: {ui_file.errorString()}")
            sys.exit(-1)
        # 从 UI 定义中动态 创建一个相应的窗口对象
        # 注意：里面的控件对象也成为窗口对象的属性了
        # 比如 self.ui.button , self.ui.textEdit
        loader = QUiLoader()
        self.ui = loader.load(ui_file)
        ui_file.close()
        if not self.ui:
            print(loader.errorString())
            sys.exit(-1)

    # 选择源文件 按钮 单击事件
    def click_selectSourceFile(self):
        global sourceFilePath, source_data
        file_name = QFileDialog.getOpenFileName(self.ui, "选择源文件", "../CylinderData",
                                                'Excel files (*.xls *.xlsx);; All files (*)')  # 选择文件，返回选中的文件路径
        print(file_name)

        if file_name:
            file_name = file_name[0]
            self.file_name = file_name
            # 根据文件类型读取数据
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                sourceFilePath = file_name
                # 读取Excel文件
                # source_data = pd.read_excel(file_name)
                # print(source_data)
                self.ui.lineEdit_sourcePath.setText(file_name)
            else:
                # 无法识别的文件类型
                print('无法识别的文件类型')
                return

    # 选择目标 按钮 单击事件
    def click_selectTargetFile(self):
        global targetFilePath
        file_name = QFileDialog.getOpenFileName(self.ui, "选择目标文件", "../CylinderData",
                                                    'Excel files (*.xls *.xlsx);; All files (*)')  # 选择文件，返回选中的文件路径
        print(file_name)

        if file_name:
            file_name = file_name[0]
            self.file_name = file_name
            # 根据文件类型读取数据
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                targetFilePath = file_name
                # 显示文件名称
                self.ui.lineEdit_targetPath.setText(file_name)
            else:
                # 无法识别的文件类型
                print('无法识别的文件类型')
                return

    #  运行 按钮 单击事件
    def click_runCopy(self):
        t_write = threading.Thread(target=write_row_by_second, )
        t_write.setDaemon(True)
        t_write.start()

    # 更新日志
    def update_log(self, msg):
        self.ui.textEdit_log.append(msg)

    def scroll_to_bottom(self):
        # 获取垂直滚动条
        scrollbar = self.ui.textEdit_log.verticalScrollBar()
        # 设置滚动条到最下面
        scrollbar.setValue(scrollbar.maximum())

def write_row_by_second():
    global sourceFilePath,  targetFilePath
    # 获取源文件的总行数
    source_workbook = openpyxl.load_workbook(sourceFilePath)
    source_sheet = source_workbook.active  # 假设源文件只有一个工作表
    total_rows = source_sheet.max_row
    # 初始化计数器
    row_to_copy = 1
    print(f"total_rows={total_rows}")
    while row_to_copy <= total_rows:
        # lock_file.acquire()  # 锁上
        # print("我正在写入，上锁")
        # 打开目标文件
        target_workbook = openpyxl.load_workbook(targetFilePath)
        target_sheet = target_workbook.active  # 假设目标文件只有一个工作表

        # 复制数据，每次复制一行,,所有列
        for col in range(1, source_sheet.max_column + 1):
            cell_value = source_sheet.cell(row=row_to_copy, column=col).value
            target_sheet.cell(row=row_to_copy, column=col, value=cell_value)

        # 保存目标文件
        target_workbook.save(targetFilePath)
        target_workbook.close()

        # 打印信息
        print(f"Copied data from row {row_to_copy}: {source_sheet[row_to_copy]}")
        msg = f"Copied data from row {row_to_copy}\n"
        window.update_log(msg=msg)
        # lock_file.release()  # 开锁
        # print("我写入完毕，开锁")
        # 增加行计数
        row_to_copy += 1

        if row_to_copy <= total_rows:
            # 1秒复制一行
            time.sleep(1)

    # 关闭源文件的工作簿
    source_workbook.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyWindow()
    window.ui.show()
    sys.exit(app.exec_())